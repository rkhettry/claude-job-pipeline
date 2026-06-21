#!/usr/bin/env python3
"""Poll the ATS boards in config/watchlist.json for new-grad SWE roles.

Fetches every resolved company's public job-board API (Ashby / Greenhouse /
Lever / SmartRecruiters), filters for new-grad-level software roles in
California or remote, and diffs against a seen-cache so only genuinely new
postings surface.

Modes:
  --test          print all current matches, do not touch the seen cache
  (default)       diff against automation/watchlist_seen.json, print only NEW
                  matches, update the cache, and append explicit new-grad
                  matches POSTED WITHIN THE LAST `--max-age` DAYS (default 2)
                  to jobs.xlsx (deduped by URL and by company+role).
                  "maybe" matches are printed only, never written.
  --max-age N     freshness window in days for xlsx writes (default 2)
  --ignore-seen   consider already-seen postings again (one-off backfills)

Posting dates come straight from the ATS APIs — the same source-of-truth data
the company careers page renders (Ashby publishedAt, Lever createdAt,
Greenhouse first_published, SmartRecruiters releasedDate). A job with no
parseable date is never written to the sheet.
"""

import json
import re
import sys
import urllib.request
import urllib.error
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timezone
from pathlib import Path

AUTOMATION_DIR = Path(__file__).resolve().parent
CONFIG = AUTOMATION_DIR / "config" / "watchlist.json"
SEEN_CACHE = AUTOMATION_DIR / "watchlist_seen.json"
UA = {"User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)"}
TIMEOUT = 10

# ---------------- filtering rules ----------------

# Titles that EXPLICITLY signal new-grad. Deliberately strict: "Engineer I",
# "Junior", "Associate" etc. do NOT count — the user only wants roles that
# literally say new grad / university grad / 0-1 years.
NEW_GRAD_PAT = re.compile(
    r"new[\s-]*grad|university\s*grad|recent\s*grad|early[\s-]*career"
    r"|entry[\s-]*level|campus\b|\b202[5-7]\b\s*(grad|start)|class\s+of\s+202[5-7]",
    re.I,
)

# Core role-shape match (from user.yaml target/high-priority role types)
ROLE_PAT = re.compile(
    r"software\s+(engineer|developer)|backend\s+engineer|full[\s-]*stack"
    r"|ml\s+engineer|machine\s+learning\s+engineer|ai\s+engineer"
    r"|applied\s+ai|member\s+of\s+technical\s+staff|forward[\s-]*deployed"
    r"|deployment\s+engineer|implementation\s+engineer|solutions\s+engineer"
    r"|infrastructure\s+engineer|platform\s+engineer|product\s+engineer"
    r"|founding\s+engineer",
    re.I,
)

# Seniority and non-SWE exclusions (from user.yaml excluded_role_levels)
EXCLUDE_PAT = re.compile(
    r"\bsenior\b|\bsr\.?\s|\bstaff\b|principal|manager|\blead\b|director"
    r"|\bvp\b|head\s+of|intern(ship)?\b|distinguished|architect"
    r"|sales\s+engineer|account\s|recruiter|designer|marketing|attorney"
    r"|counsel|finance|accountant|\bIT\b\s+support",
    re.I,
)

# JD-text signals — also strict: explicit new-grad language or 0-1 years only.
# "0-2 years" / generic "early career" in a JD body no longer qualifies.
JD_NEWGRAD_PAT = re.compile(
    r"new[\s-]*grad|recent\s+graduate|university\s+graduate"
    r"|0\s*(?:[-–]|to)\s*1\s+year|zero\s+to\s+one\s+year"
    r"|less\s+than\s+(?:one|1)\s+year\s+of\s+experience"
    r"|graduating\s+in\s+202[5-7]",
    re.I,
)

CA_LOCATION_PAT = re.compile(
    r"san\s*francisco|\bsf\b|bay\s*area|palo\s*alto|mountain\s*view|san\s*jose"
    r"|sunnyvale|santa\s*clara|menlo\s*park|redwood\s*city|san\s*mateo"
    r"|foster\s*city|oakland|berkeley|burlingame|south\s*san\s*francisco"
    r"|los\s*gatos|fremont|cupertino|emeryville|san\s*bruno"
    r"|los\s*angeles|\bla\b|santa\s*monica|el\s*segundo|torrance|long\s*beach"
    r"|hawthorne|costa\s*mesa|irvine|glendale|culver\s*city|venice"
    r"|san\s*diego|california|\bca\b(?![a-z])",
    re.I,
)
REMOTE_PAT = re.compile(r"remote", re.I)


NON_US_PAT = re.compile(
    r"spain|poland|singapore|london|\buk\b|united\s+kingdom|canada|toronto"
    r"|vancouver|india|bangalore|bengaluru|hyderabad|germany|berlin|munich"
    r"|france|paris|netherlands|amsterdam|ireland|dublin|australia|sydney"
    r"|japan|tokyo|brazil|mexico\b|israel|tel\s*aviv|portugal|lisbon"
    r"|emea|apac|latam",
    re.I,
)


def location_ok(loc):
    if not loc:
        return False  # unknown location: skip rather than spam
    if CA_LOCATION_PAT.search(loc):
        return True
    # remote counts only when it isn't a non-US remote market
    return bool(REMOTE_PAT.search(loc) and not NON_US_PAT.search(loc))


def classify(title, jd_text=""):
    """Return 'new-grad', 'maybe' (role matches + JD hints ok), or None."""
    if EXCLUDE_PAT.search(title):
        return None
    # must be a software-shaped role either way — otherwise "Early Career
    # Mechanical Engineer" / "Propulsion Engineer I" style titles flood in
    if not ROLE_PAT.search(title):
        return None
    if NEW_GRAD_PAT.search(title):
        return "new-grad"
    if jd_text and JD_NEWGRAD_PAT.search(jd_text):
        return "new-grad"
    return "maybe"


# ---------------- per-ATS fetchers ----------------
# Each returns a list of dicts: {id, title, location, url, posted, jd_text}

def fetch_json(url, _retry=True):
    req = urllib.request.Request(url, headers=UA)
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT) as resp:
            return json.loads(resp.read().decode("utf-8", errors="replace"))
    except Exception:
        if _retry:
            import time
            time.sleep(1.5)
            return fetch_json(url, _retry=False)
        return None


def strip_html(html):
    return re.sub(r"<[^>]+>", " ", html or "")[:4000]


def jobs_ashby(slug):
    data = fetch_json(f"https://api.ashbyhq.com/posting-api/job-board/{slug}?includeCompensation=true")
    out = []
    for j in (data or {}).get("jobs", []):
        locs = [j.get("location") or ""] + [s.get("location", "") for s in j.get("secondaryLocations", [])]
        if j.get("isRemote"):
            locs.append("Remote")
        comp = (j.get("compensation") or {}).get("compensationTierSummary", "")
        out.append({
            "id": f"ashby/{slug}/{j.get('id')}",
            "title": j.get("title", ""),
            "location": "; ".join(l for l in locs if l),
            "url": j.get("jobUrl") or j.get("applyUrl") or "",
            "posted": (j.get("publishedAt") or "")[:10],
            "jd_text": strip_html(j.get("descriptionHtml", "")),
            "comp": comp,
        })
    return out


def jobs_greenhouse(slug):
    data = fetch_json(f"https://boards-api.greenhouse.io/v1/boards/{slug}/jobs?content=true")
    out = []
    for j in (data or {}).get("jobs", []):
        out.append({
            "id": f"greenhouse/{slug}/{j.get('id')}",
            "title": j.get("title", ""),
            "location": (j.get("location") or {}).get("name", ""),
            "url": j.get("absolute_url", ""),
            # first_published ONLY — updated_at changes on any edit and makes
            # months-old postings look fresh
            "posted": (j.get("first_published") or "")[:10],
            "jd_text": strip_html(j.get("content", "")),
            "comp": "",
        })
    return out


def jobs_lever(slug):
    data = fetch_json(f"https://api.lever.co/v0/postings/{slug}?mode=json")
    out = []
    for j in (data or []):
        cat = j.get("categories") or {}
        ts = j.get("createdAt")
        posted = datetime.fromtimestamp(ts / 1000, tz=timezone.utc).strftime("%Y-%m-%d") if ts else ""
        loc = cat.get("location", "") or ""
        if j.get("workplaceType") == "remote":
            loc += "; Remote"
        out.append({
            "id": f"lever/{slug}/{j.get('id')}",
            "title": j.get("text", ""),
            "location": loc,
            "url": j.get("hostedUrl", ""),
            "posted": posted,
            "jd_text": (j.get("descriptionPlain") or "")[:4000],
            "comp": "",
        })
    return out


def jobs_smartrecruiters(slug):
    data = fetch_json(f"https://api.smartrecruiters.com/v1/companies/{slug}/postings?limit=100")
    out = []
    for j in (data or {}).get("content", []):
        loc = j.get("location") or {}
        loc_str = ", ".join(x for x in [loc.get("city", ""), loc.get("region", "")] if x)
        if loc.get("remote"):
            loc_str += "; Remote"
        out.append({
            "id": f"smartrecruiters/{slug}/{j.get('id')}",
            "title": j.get("name", ""),
            "location": loc_str,
            "url": f"https://jobs.smartrecruiters.com/{slug}/{j.get('id')}",
            "posted": (j.get("releasedDate") or "")[:10],
            "jd_text": "",  # needs a per-job fetch; title-level filtering only
            "comp": "",
        })
    return out


FETCHERS = {
    "ashby": jobs_ashby,
    "greenhouse": jobs_greenhouse,
    "lever": jobs_lever,
    "smartrecruiters": jobs_smartrecruiters,
}


def poll_company(company):
    fetcher = FETCHERS.get(company.get("ats"))
    if not fetcher:
        return company, []
    jobs = fetcher(company["slug"]) or []
    matches = []
    for j in jobs:
        kind = classify(j["title"], j["jd_text"])
        if kind and location_ok(j["location"]):
            j["match"] = kind
            j["company"] = company["name"]
            j["category"] = company["category"]
            matches.append(j)
    return company, matches


def fresh_enough(m, max_age_days):
    """True only when the posting has a real date within the window."""
    try:
        age = (datetime.now() - datetime.strptime(m.get("posted") or "", "%Y-%m-%d")).days
    except ValueError:
        return False
    return 0 <= age <= max_age_days


def main():
    test_mode = "--test" in sys.argv
    ignore_seen = "--ignore-seen" in sys.argv
    max_age = 2
    if "--max-age" in sys.argv:
        max_age = int(sys.argv[sys.argv.index("--max-age") + 1])
    cfg = json.loads(CONFIG.read_text())
    companies, seen_boards = [], set()
    for c in cfg["companies"]:
        if c.get("ats") not in FETCHERS:
            continue
        board = (c["ats"], c["slug"])  # two entries can resolve to one board
        if board in seen_boards:
            continue
        seen_boards.add(board)
        companies.append(c)
    print(f"Polling {len(companies)} company boards...")

    all_matches, failures = [], []
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(poll_company, c): c["name"] for c in companies}
        for fut in as_completed(futures):
            company, matches = fut.result()
            all_matches.extend(matches)

    seen = set()
    if SEEN_CACHE.exists():
        seen = set(json.loads(SEEN_CACHE.read_text()).get("ids", []))

    new = [m for m in all_matches if ignore_seen or m["id"] not in seen]
    shown = all_matches if test_mode else new

    shown.sort(key=lambda m: (m["match"] != "new-grad", m.get("posted") or ""), reverse=False)
    ng = [m for m in shown if m["match"] == "new-grad"]
    mb = [m for m in shown if m["match"] == "maybe"]

    print(f"\n{'='*100}")
    print(f"EXPLICIT NEW-GRAD ROLES ({len(ng)}):\n")
    for m in ng:
        comp = f"  [{m['comp']}]" if m.get("comp") else ""
        print(f"  {m['posted'] or '????-??-??'}  {m['company']:<22} {m['title'][:58]:<58}{comp}")
        print(f"              {m['location'][:70]}")
        print(f"              {m['url']}")
    print(f"\n{'-'*100}")
    print(f"POSSIBLE FITS — right role shape, level not explicit ({len(mb)}):\n")
    for m in mb[:40]:
        print(f"  {m['posted'] or '????-??-??'}  {m['company']:<22} {m['title'][:70]}")
    if len(mb) > 40:
        print(f"  ... and {len(mb) - 40} more")

    if not test_mode:
        writable = [m for m in new if m["match"] == "new-grad" and fresh_enough(m, max_age)]
        skipped_stale = [m for m in new if m["match"] == "new-grad" and not fresh_enough(m, max_age)]
        added = append_to_xlsx(writable)
        seen.update(m["id"] for m in all_matches)
        SEEN_CACHE.write_text(json.dumps({"ids": sorted(seen)}, indent=0) + "\n")
        print(f"\nSeen cache updated: {len(seen)} ids tracked. New this run: {len(new)}. "
              f"Added to jobs.xlsx: {added} (posted within {max_age}d); "
              f"skipped as stale/undated: {len(skipped_stale)}")


def _norm(s):
    return re.sub(r"[^a-z0-9]", "", (s or "").lower())


def append_to_xlsx(matches):
    """Append new-grad matches to jobs.xlsx, deduped against existing rows."""
    if not matches:
        return 0
    import openpyxl
    xlsx = AUTOMATION_DIR / "jobs.xlsx"
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["jobs"]
    headers = [c.value for c in ws[1]]
    col = {h: i for i, h in enumerate(headers)}

    existing_urls, existing_pairs, max_id = set(), set(), 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        existing_urls.add((row[col["apply_url"]] or "").split("?")[0].rstrip("/"))
        existing_pairs.add((_norm(row[col["company"]]), _norm(row[col["role"]])))
        try:
            max_id = max(max_id, int(row[col["id"]]))
        except (TypeError, ValueError):
            pass

    today = datetime.now().strftime("%Y-%m-%d")
    added = 0
    for m in matches:
        url_key = m["url"].split("?")[0].rstrip("/")
        pair = (_norm(m["company"]), _norm(m["title"]))
        if url_key in existing_urls or pair in existing_pairs:
            continue
        try:
            age = (datetime.now() - datetime.strptime(m["posted"], "%Y-%m-%d")).days
        except ValueError:
            age = 99
        tier = "A" if m["category"] in (
            "frontier-lab", "ai-agents", "ai-infra", "voice-ai", "defense-aero") else "B"
        rowvals = {
            "id": max_id + added + 1,
            "date_sourced": today,
            "posted_date": m["posted"],
            "freshness": "recent" if age <= 3 else "older",
            "company": m["company"],
            "role": m["title"].strip(),
            "location": m["location"][:120],
            "comp": m.get("comp", ""),
            "source": "watchlist",
            "apply_url": m["url"],
            "tier": tier,
            "reasoning": f"watchlist direct-from-ATS hit ({m['category']}); "
                         f"explicit new-grad/entry title or JD match",
            "decision": "",
            "jd_text": m["jd_text"][:2000],
        }
        ws.append([rowvals.get(h, "") for h in headers])
        existing_urls.add(url_key)
        existing_pairs.add(pair)
        added += 1
    if added:
        wb.save(xlsx)
    return added


if __name__ == "__main__":
    main()
