#!/usr/bin/env python3
"""Localhost web UI for triaging jobs in <repo>/automation/jobs.xlsx, plus a live
pipeline view (`/runs`) that visualizes the daily sourcing run.

Run from terminal (cd into your clone first):
    python3 ./automation/triage_server.py

Then open http://localhost:8765/ in your browser.
Press Ctrl+C in the terminal to stop.

Identity, preferences, and prestige bar live in automation/config/user.yaml.
Edit that file, not this server.
"""

import json
import os
import re
import shlex
import sys
import shutil
import subprocess
import webbrowser
import threading
import datetime
import mimetypes
from urllib.parse import unquote
from http.server import HTTPServer, BaseHTTPRequestHandler
from pathlib import Path

# Auto-install openpyxl if missing (one-time)
try:
    from openpyxl import load_workbook
except ImportError:
    print("Installing openpyxl (one-time)...", file=sys.stderr)
    subprocess.run(
        [sys.executable, "-m", "pip", "install", "openpyxl", "--quiet", "--break-system-packages"],
        check=False,
    )
    from openpyxl import load_workbook  # type: ignore

# Paths
AUTOMATION_DIR = Path(__file__).resolve().parent
# RESUME_ROOT defaults to the repo root (parent of automation/). Override with the
# RESUME_ROOT env var if you keep your resume materials in a separate location.
RESUME_ROOT = Path(os.environ.get("RESUME_ROOT", str(AUTOMATION_DIR.parent)))
USER_CONFIG = AUTOMATION_DIR / "config" / "user.yaml"
XLSX_PATH = AUTOMATION_DIR / "jobs.xlsx"
SPEC_PATH = AUTOMATION_DIR / "resume-tailoring-spec.md"
LOGS_DIR = AUTOMATION_DIR / "logs"
LOGS_DIR.mkdir(exist_ok=True)
OUTREACH_DIR = AUTOMATION_DIR / "outreach"
OUTREACH_DIR.mkdir(exist_ok=True)
OUTREACH_FIND_PROMPT = AUTOMATION_DIR / "prompts" / "outreach-find.txt"
OUTREACH_SEND_PROMPT = AUTOMATION_DIR / "prompts" / "outreach-send.txt"
VERIFY_DATE_PROMPT = AUTOMATION_DIR / "prompts" / "verify-date.txt"
PORT = int(os.environ.get("TRIAGE_PORT", "8765"))

# Sourcing pipeline stages — used by /runs live-view to show which step Claude is on.
# Order is the pipeline order (preflight → ... → done). Patterns matched against
# ANSI-stripped log content; the LATEST stage seen wins as "current".
STAGE_DEFS = [
    ("preflight",          "Pre-flight",        r"(?i)(read(ing)?\s+(the\s+)?spec|read(ing)?\s+.*jobs\.xlsx|existing.*rows|build.*dedup|pre.?flight)"),
    ("simplify",           "SimplifyJobs",      r"(?i)(simplifyjobs|simplify[\s_\-]*jobs|simplify.*readme)"),
    ("jobright_repo",      "jobright-ai repo",  r"(?i)(jobright-ai/2026|jobright.*repo|2026-software-engineer)"),
    ("jobright_minisite",  "jobright minisite", r"(?i)(jobright\.ai/minisites|newgrad-jobs|is\s+new\s+grad|jobright.*minisite|claude-in-chrome.*jobright)"),
    ("hn_hiring",          "HN Who's Hiring",   r"(?i)(news\.ycombinator|hn\.algolia|who.{0,2}s\s+hiring|hn[\s_-]+hiring|ask\s+hn.*hiring|hiring\?.*\(.*20\d\d\))"),
    ("jd_verify",          "JD verify",         r"(?i)(jd[\s_-]*verif|job description|read(ing)?.*apply.url|verify.*new.?grad|1\+\s*year|2\+\s*year)"),
    ("filter",             "Filter & prestige", r"(?i)(apply.*prestige|prestige bar|hard.?avoid|tier[\s\-]*a\b)"),
    ("dedupe",             "Dedupe",            r"(?i)(dedup|canonical.*url|already in sheet|skip.*dup)"),
    ("append",             "Append rows",       r"(?i)(openpyxl|wb\.save|append.*row|appending|writing.*xlsx)"),
    ("email",              "Email digest",      r"(?i)(mail\.google\.com|gmail|compose|digest email)"),
    ("done",               "DONE",              r"(?i)(done\s+[—\-].+added|^done\s+[—\-]|finished.*one-line)"),
]
ANSI_RE = re.compile(r"\x1b\[[0-9;]*[a-zA-Z]")


def _strip_ansi(s: str) -> str:
    return ANSI_RE.sub("", s)


def _find_active_sourcing_log():
    """Return the most-recent sourcing-*.log (Path) or None."""
    candidates = sorted(
        LOGS_DIR.glob("sourcing-*.log"),
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def _is_run_active(log_path) -> bool:
    """A run is 'active' if the log file was modified in the last 90 seconds."""
    if not log_path or not log_path.exists():
        return False
    age = datetime.datetime.now().timestamp() - log_path.stat().st_mtime
    return age < 90


def _detect_stages(log_text: str):
    """Walk the ANSI-stripped log; return (current_id, set(completed_ids)).

    Current = highest-index stage that has matched anywhere in the log.
    Completed = every stage that matched somewhere, minus current.
    """
    seen = set()
    max_idx = -1
    for i, (sid, _, pattern) in enumerate(STAGE_DEFS):
        if re.search(pattern, log_text):
            seen.add(sid)
            if i > max_idx:
                max_idx = i
    if max_idx < 0:
        return None, set()
    current = STAGE_DEFS[max_idx][0]
    return current, seen - {current}

# Claude Code invocation
CLAUDE_BIN = shutil.which("claude") or "claude"


def spawn_claude_tailor(job_id: int) -> Path:
    """Open Terminal.app running Claude Code interactively with the tailoring prompt.

    The prompt is passed as a CLI argument to `claude`, NOT typed in after launch —
    this eliminates the keystroke-injection race that left Claude sitting idle when
    the typing happened before the TUI was ready to accept input.

    Returns the path to the launch log file.
    """
    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    log_path = LOGS_DIR / f"tailor-{job_id}-{timestamp}.log"

    prompt = (
        f"Read {SPEC_PATH} and execute it for job_id={job_id}. "
        f"The job row lives at id={job_id} in {XLSX_PATH}. "
        f"Working directory is {RESUME_ROOT}. "
        f"**If the row has a `jd_text` column populated, use that as the JD source directly — "
        f"skip web_fetching the apply_url.** That field is set when the user manually pastes a "
        f"JD via the triage UI's Add Job form. "
        f"**Pre-step (if applicable):** if the row's `company` or `role` is empty or equals "
        f"'(extracting…)', figure out the real values BEFORE tailoring. (a) If `jd_text` is "
        f"populated, extract company / role / location from it. (b) Otherwise if `apply_url` is "
        f"populated, fetch that URL (Claude in Chrome for JS-rendered ATS like Workday/Lever/"
        f"Ashby; `web_fetch` for static pages) to get the JD, then extract. Write the extracted "
        f"`company`, `role`, and (if clearly stated) `location` back to jobs.xlsx for this "
        f"job_id via openpyxl. If you fetched the URL and got the JD body, also write it to the "
        f"`jd_text` column so future re-tailoring doesn't have to re-fetch. Then proceed with "
        f"the §5 verification loop. "
        f"Follow the agentic verification loop in §5 of the spec (max 4 iterations). "
        f"When done, update the resume_version column in jobs.xlsx for this job_id per §7 "
        f"(use openpyxl from Bash). "
        f"If anything fails, write 'error: <one-line reason>' to that column instead. "
        f"Begin now."
    )

    # Per-spawn close helper. Sets a unique window title via the OSC 0 escape,
    # then closes whichever Terminal window has that exact title once claude exits.
    # Lives next to the log so cleanup is easy + the file is unique per spawn.
    window_title = f"claude-tailor-{job_id}-{timestamp}"
    close_script = LOGS_DIR / f"close-{job_id}-{timestamp}.sh"
    close_script.write_text(
        "#!/bin/bash\n"
        f"osascript -e 'tell application \"Terminal\" to close "
        f"(every window whose name contains \"{window_title}\") saving no' "
        "2>/dev/null\n"
    )
    close_script.chmod(0o755)

    # Build the shell command. shlex.quote handles all shell escaping (the prompt
    # may contain $, quotes, etc). `-p` is non-interactive print mode: claude does
    # the task and EXITS (vs. interactive mode which sits at a prompt forever).
    # That makes the close-helper fire after claude exits, so the Terminal window
    # disposes of itself automatically. Without -p, the window stays open until
    # the user manually /exits.
    # stream-json (+ required --verbose) emits one JSON event per line AS THE RUN
    # HAPPENS, teed into the log — that's what feeds the 📡 live viewer at /runs.
    # Plain -p would buffer everything and keep the log empty until the very end.
    inner_cmd = (
        f"printf '\\033]0;{window_title}\\007'; "
        f"cd {shlex.quote(str(RESUME_ROOT))} && "
        f"claude -p --verbose --output-format stream-json --dangerously-skip-permissions "
        f"{shlex.quote(prompt)} 2>&1 | tee -a {shlex.quote(str(log_path))}; "
        f"bash {shlex.quote(str(close_script))} & "
        f"sleep 1; exit"
    )

    # AppleScript-escape the inner command for embedding in `do script "..."`
    cmd_esc = inner_cmd.replace("\\", "\\\\").replace('"', '\\"')

    # Use Terminal.app only — the user has a custom iTerm config that breaks keystroke
    # injection. With the prompt passed as CLI arg we don't need keystroke injection
    # anyway, but Terminal.app is more reliable end-to-end.
    applescript = f'''
tell application "Terminal"
    activate
    do script "{cmd_esc}"
end tell
'''

    script_path = LOGS_DIR / f"launch-{job_id}-{timestamp}.applescript"
    script_path.write_text(applescript)

    log_fh = open(log_path, "w", encoding="utf-8", errors="replace")
    log_fh.write(f"=== tailor launch job_id={job_id} at {timestamp} ===\n")
    log_fh.write(f"cwd: {RESUME_ROOT}\n")
    log_fh.write(f"prompt: {prompt}\n")
    log_fh.write(f"applescript: {script_path}\n")
    log_fh.write("=" * 60 + "\n")

    try:
        result = subprocess.run(
            ["osascript", str(script_path)],
            capture_output=True,
            text=True,
            timeout=20,
        )
        log_fh.write(f"osascript exit: {result.returncode}\n")
        if result.stdout:
            log_fh.write(f"stdout: {result.stdout}\n")
        if result.stderr:
            log_fh.write(f"stderr: {result.stderr}\n")
        if result.returncode != 0:
            _set_resume_version(
                job_id,
                f"error: terminal launch failed (osascript exit {result.returncode})",
            )
    except FileNotFoundError:
        log_fh.write("ERROR: osascript not found — are you on macOS?\n")
        _set_resume_version(job_id, "error: osascript missing (macOS required)")
    except subprocess.TimeoutExpired:
        log_fh.write("ERROR: osascript timed out launching terminal.\n")
        _set_resume_version(job_id, "error: terminal launch timed out")
    except Exception as e:
        log_fh.write(f"ERROR: {e}\n")
        _set_resume_version(job_id, f"error: terminal launch raised {type(e).__name__}")
    finally:
        log_fh.close()

    return log_path


def _set_resume_version(job_id: int, value: str) -> None:
    """Write to the resume_version cell for a given job_id."""
    try:
        wb = load_workbook(XLSX_PATH)
        ws = wb["jobs"]
        headers = [c.value for c in ws[1]]
        if "resume_version" not in headers:
            return
        rv_idx = headers.index("resume_version")
        for r in ws.iter_rows(min_row=2):
            if r[0].value == job_id:
                r[rv_idx].value = value
                break
        wb.save(XLSX_PATH)
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Outreach (LinkedIn lead-finding + sending). Two Claude Code spawns: one to
# discover recruiters/founders/engineers, one to actually send the messages.
# Both use Chrome MCP (the user is already authenticated on LinkedIn).
# ---------------------------------------------------------------------------

def _outreach_sidecar(job_id: int) -> Path:
    return OUTREACH_DIR / f"{job_id}.json"


def _spawn_claude_outreach(job_id: int, kind: str, lead_count: int = 0, append: bool = False) -> Path:
    """Open Terminal.app running Claude Code with the outreach prompt.

    kind: "find" (Stage 1, discover leads) or "send" (Stage 2, send messages).
    Returns the path to the launch log file.

    Pattern matches spawn_claude_tailor: prompt as CLI arg + per-spawn close helper
    so the Terminal window self-destructs once claude exits.
    """
    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    log_path = LOGS_DIR / f"outreach-{kind}-{job_id}-{timestamp}.log"

    if kind == "find":
        if not OUTREACH_FIND_PROMPT.exists():
            raise FileNotFoundError(f"outreach-find prompt missing at {OUTREACH_FIND_PROMPT}")
        template = OUTREACH_FIND_PROMPT.read_text()
        prompt = (template
                  .replace("{JOB_ID}", str(job_id))
                  .replace("{LEAD_COUNT}", str(lead_count))
                  .replace("{APPEND}", "true" if append else "false"))
    elif kind == "send":
        if not OUTREACH_SEND_PROMPT.exists():
            raise FileNotFoundError(f"outreach-send prompt missing at {OUTREACH_SEND_PROMPT}")
        template = OUTREACH_SEND_PROMPT.read_text()
        prompt = template.replace("{JOB_ID}", str(job_id))
    else:
        raise ValueError(f"unknown outreach kind: {kind}")

    window_title = f"claude-outreach-{kind}-{job_id}-{timestamp}"
    close_script = LOGS_DIR / f"close-outreach-{kind}-{job_id}-{timestamp}.sh"
    close_script.write_text(
        "#!/bin/bash\n"
        f"osascript -e 'tell application \"Terminal\" to close "
        f"(every window whose name contains \"{window_title}\") saving no' "
        "2>/dev/null\n"
    )
    close_script.chmod(0o755)

    # --chrome flag enables Chrome MCP so Claude can drive LinkedIn directly.
    # `-p` is non-interactive: claude exits after the task so the close-helper fires.
    # stream-json + tee feeds the 📡 live viewer at /runs (see tailor spawn).
    inner_cmd = (
        f"printf '\\033]0;{window_title}\\007'; "
        f"cd {shlex.quote(str(RESUME_ROOT))} && "
        f"claude -p --verbose --output-format stream-json --dangerously-skip-permissions "
        f"--chrome {shlex.quote(prompt)} 2>&1 | tee -a {shlex.quote(str(log_path))}; "
        f"bash {shlex.quote(str(close_script))} & "
        f"sleep 1; exit"
    )
    cmd_esc = inner_cmd.replace("\\", "\\\\").replace('"', '\\"')
    applescript = f'''
tell application "Terminal"
    activate
    do script "{cmd_esc}"
end tell
'''
    script_path = LOGS_DIR / f"launch-outreach-{kind}-{job_id}-{timestamp}.applescript"
    script_path.write_text(applescript)

    log_fh = open(log_path, "w", encoding="utf-8", errors="replace")
    log_fh.write(f"=== outreach-{kind} launch job_id={job_id} at {timestamp} ===\n")
    log_fh.write(f"cwd: {RESUME_ROOT}\n")
    log_fh.write(f"prompt:\n{prompt}\n")
    log_fh.write(f"applescript: {script_path}\n")
    log_fh.write("=" * 60 + "\n")

    try:
        result = subprocess.run(
            ["osascript", str(script_path)],
            capture_output=True, text=True, timeout=20,
        )
        log_fh.write(f"osascript exit: {result.returncode}\n")
        if result.stdout: log_fh.write(f"stdout: {result.stdout}\n")
        if result.stderr: log_fh.write(f"stderr: {result.stderr}\n")
    except Exception as e:
        log_fh.write(f"ERROR: {e}\n")
    finally:
        log_fh.close()

    return log_path


def _load_sidecar(job_id: int):
    """Return the parsed sidecar JSON for a job, or None if missing/invalid."""
    sidecar = _outreach_sidecar(job_id)
    if not sidecar.exists():
        return None
    try:
        return json.loads(sidecar.read_text())
    except Exception:
        return None


# ---------------------------------------------------------------------------
# Terminal cleanup. Targets Terminal.app ONLY (never iTerm2). Window titles all
# start with "claude-" (claude-tailor-, claude-outreach-, claude-verify-date-,
# claude-sourcing-). Closing a window cascades SIGHUP to the shell and any
# claude process inside, so window-close is usually sufficient. We also pkill
# orphans (claudes whose Terminal window was already manually closed) to keep
# the system clean.
# ---------------------------------------------------------------------------

# Spec-file patterns that uniquely identify task-spawn claude processes. These
# never match the user's interactive iTerm2 claude sessions (no spec file in
# their cmdline).
SPAWN_SPEC_PATTERNS = [
    "resume-tailoring-spec.md",
    "outreach-lead-finder-spec.md",
    "outreach-sender-spec.md",
    "verify-date-spec.md",
    "job-sourcing-spec.md",
]


def _close_terminal_windows(title_substring: str) -> int:
    """Close every Terminal.app (NOT iTerm2) window whose name contains the substring.

    Returns the number of windows closed (best-effort; osascript success).
    Empty substring would close everything matching "claude-" prefix; never accept that.
    """
    if not title_substring or "claude-" not in title_substring:
        # Safety: refuse to close everything. Always scoped to claude-* titles.
        return 0
    script = (
        f'tell application "Terminal"\n'
        f'  set n to count (every window whose name contains "{title_substring}")\n'
        f'  close (every window whose name contains "{title_substring}") saving no\n'
        f'  return n\n'
        f'end tell\n'
    )
    try:
        result = subprocess.run(
            ["osascript", "-e", script],
            capture_output=True, text=True, timeout=10,
        )
        if result.returncode == 0 and result.stdout.strip().isdigit():
            return int(result.stdout.strip())
    except Exception:
        pass
    return 0


def _kill_orphan_claude_procs(job_id: int | None = None) -> int:
    """Kill task-spawn claude processes (orphans or otherwise).

    If job_id is given, restrict to processes whose cmdline references that id.
    Otherwise kill all task-spawn claudes (those running one of SPAWN_SPEC_PATTERNS).

    Returns approximate count killed.

    Implementation note: the task prompts have NEWLINES between "claude" and the
    spec file path (multi-line prompt embedded in argv). pgrep -f's `.` doesn't
    cross newlines, so a pattern like `claude.*outreach-sender-spec.md` won't
    match. We pgrep by the spec filename ALONE (uniquely identifies a task spawn
    in practice; no other process embeds these filenames in its cmdline), then
    confirm each candidate PID is actually a `claude` process via ps before kill.
    """
    killed = 0
    seen = set()
    for spec in SPAWN_SPEC_PATTERNS:
        try:
            result = subprocess.run(
                ["pgrep", "-f", spec],
                capture_output=True, text=True, timeout=5,
            )
        except Exception:
            continue
        for pid in result.stdout.split():
            pid = pid.strip()
            if not pid or pid in seen:
                continue
            seen.add(pid)
            # Confirm this is a `claude` process (not e.g. an editor with the
            # file open). ps -o comm prints the binary name.
            try:
                p = subprocess.run(
                    ["ps", "-p", pid, "-o", "comm="],
                    capture_output=True, text=True, timeout=2,
                )
                if "claude" not in p.stdout.lower():
                    continue
            except Exception:
                continue
            # If job_id constraint set, confirm this process references that id.
            if job_id is not None:
                try:
                    full = subprocess.run(
                        ["ps", "-p", pid, "-o", "command="],
                        capture_output=True, text=True, timeout=2,
                    ).stdout
                    if not (
                        f"JOB_ID = {job_id}" in full
                        or f"job_id={job_id}" in full
                        or f"job_id = {job_id}" in full
                    ):
                        continue
                except Exception:
                    continue
            try:
                subprocess.run(["kill", pid], timeout=2)
                killed += 1
            except Exception:
                pass
    return killed


def _reset_inflight_state(job_id: int | None = None) -> dict:
    """Reset xlsx + sidecar fields that indicate a task is in-flight, so the UI
    returns to a usable state after a kill. Returns counts of what was reset.

    If job_id is given, only that row. Otherwise scan all rows.
    """
    counts = {"resume_version": 0, "posted_date_verified": 0, "outreach_stage": 0}
    try:
        wb = load_workbook(XLSX_PATH)
    except Exception:
        return counts
    ws = wb["jobs"]
    headers = [c.value for c in ws[1]]
    rv_idx = headers.index("resume_version") if "resume_version" in headers else None
    pdv_idx = headers.index("posted_date_verified") if "posted_date_verified" in headers else None
    decision_idx = headers.index("decision") if "decision" in headers else None
    changed = False
    affected_ids = []
    for r in ws.iter_rows(min_row=2):
        rid = r[0].value
        if rid is None:
            continue
        if job_id is not None and rid != job_id:
            continue
        # Reset tailoring spinner → revert decision to pending
        if rv_idx is not None and (r[rv_idx].value or "") == "tailoring":
            r[rv_idx].value = ""
            if decision_idx is not None:
                r[decision_idx].value = "pending"
            counts["resume_version"] += 1
            changed = True
            affected_ids.append(rid)
        # Reset verify-date spinner
        if pdv_idx is not None and (r[pdv_idx].value or "") == "verifying":
            r[pdv_idx].value = ""
            counts["posted_date_verified"] += 1
            changed = True
            affected_ids.append(rid)
    if changed:
        wb.save(XLSX_PATH)

    # Reset outreach sidecars (stage = finding / sending → leads-ready)
    sidecar_ids = [job_id] if job_id is not None else None
    sidecars = (
        [_outreach_sidecar(job_id)] if job_id is not None
        else list(OUTREACH_DIR.glob("*.json"))
    )
    for s in sidecars:
        try:
            d = json.loads(s.read_text())
        except Exception:
            continue
        if d.get("stage") in ("finding", "sending"):
            d["stage"] = "leads-ready"
            s.write_text(json.dumps(d, indent=2, ensure_ascii=False))
            counts["outreach_stage"] += 1
    return counts


def _spawn_claude_verify_date(job_id: int) -> Path:
    """Spawn Claude Code with --chrome to verify a job's posted date from the source ATS page.

    Writes the result to the `posted_date_verified` column in jobs.xlsx.
    Returns the path to the launch log file. Pattern matches the other spawners
    (Terminal.app + prompt as CLI arg + self-closing window).
    """
    timestamp = datetime.datetime.now().strftime("%Y%m%d-%H%M%S")
    log_path = LOGS_DIR / f"verify-date-{job_id}-{timestamp}.log"

    if not VERIFY_DATE_PROMPT.exists():
        raise FileNotFoundError(f"verify-date prompt missing at {VERIFY_DATE_PROMPT}")
    prompt = VERIFY_DATE_PROMPT.read_text().replace("{JOB_ID}", str(job_id))

    window_title = f"claude-verify-date-{job_id}-{timestamp}"
    close_script = LOGS_DIR / f"close-verify-date-{job_id}-{timestamp}.sh"
    close_script.write_text(
        "#!/bin/bash\n"
        f"osascript -e 'tell application \"Terminal\" to close "
        f"(every window whose name contains \"{window_title}\") saving no' "
        "2>/dev/null\n"
    )
    close_script.chmod(0o755)

    # `-p` is non-interactive print mode: claude exits after the task so the
    # close-helper fires and the Terminal window auto-disposes.
    # stream-json + tee feeds the 📡 live viewer at /runs (see tailor spawn).
    inner_cmd = (
        f"printf '\\033]0;{window_title}\\007'; "
        f"cd {shlex.quote(str(RESUME_ROOT))} && "
        f"claude -p --verbose --output-format stream-json --dangerously-skip-permissions "
        f"--chrome {shlex.quote(prompt)} 2>&1 | tee -a {shlex.quote(str(log_path))}; "
        f"bash {shlex.quote(str(close_script))} & "
        f"sleep 1; exit"
    )
    cmd_esc = inner_cmd.replace("\\", "\\\\").replace('"', '\\"')
    applescript = f'''
tell application "Terminal"
    activate
    do script "{cmd_esc}"
end tell
'''
    script_path = LOGS_DIR / f"launch-verify-date-{job_id}-{timestamp}.applescript"
    script_path.write_text(applescript)

    with open(log_path, "w", encoding="utf-8", errors="replace") as log_fh:
        log_fh.write(f"=== verify-date launch job_id={job_id} at {timestamp} ===\n")
        log_fh.write(f"prompt:\n{prompt}\n{'='*60}\n")
        try:
            result = subprocess.run(
                ["osascript", str(script_path)],
                capture_output=True, text=True, timeout=20,
            )
            log_fh.write(f"osascript exit: {result.returncode}\n")
            if result.stdout: log_fh.write(f"stdout: {result.stdout}\n")
            if result.stderr: log_fh.write(f"stderr: {result.stderr}\n")
        except Exception as e:
            log_fh.write(f"ERROR: {e}\n")

    return log_path


def _set_posted_date_verified(job_id: int, value: str) -> bool:
    """Write to (or migrate-then-write) the posted_date_verified column.

    Returns True if the row was found and updated, False if the job_id doesn't exist.
    The 'verifying' marker is used as a transient state so the UI shows a spinner.
    """
    try:
        wb = load_workbook(XLSX_PATH)
    except PermissionError:
        return False
    ws = wb["jobs"]
    headers = [c.value for c in ws[1]]
    # Migrate column if missing — same pattern as jd_text. Plain text header, no
    # font copy to avoid the StyleProxy unhashable error on wb.save().
    if "posted_date_verified" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="posted_date_verified")
        headers.append("posted_date_verified")
    idx = headers.index("posted_date_verified")
    found = False
    for r in ws.iter_rows(min_row=2):
        if r[0].value == job_id:
            r[idx].value = value
            found = True
            break
    if found:
        wb.save(XLSX_PATH)
    return found


def _outreach_summary(job_id: int):
    """Compact dict for /api/jobs to inject into each row.

    Returns {} if no sidecar exists. Otherwise computes counts + a status
    string the UI uses to decide which button to show.
    """
    data = _load_sidecar(job_id)
    if not data:
        return {}
    leads = data.get("leads", []) or []
    sent = sum(1 for L in leads if L.get("send_status") == "sent")
    failed = sum(1 for L in leads if L.get("send_status") == "failed")
    approved = sum(1 for L in leads if L.get("approved") is True)
    pending_review = sum(1 for L in leads if L.get("approved") is None)
    return {
        "outreach_stage": data.get("stage", "unknown"),
        "outreach_lead_count": len(leads),
        "outreach_sent": sent,
        "outreach_failed": failed,
        "outreach_approved": approved,
        "outreach_pending_review": pending_review,
        "outreach_found_at": data.get("found_at"),
        "outreach_sent_at": data.get("sent_at"),
    }

# ---------------------------------------------------------------------------
# HTML/JS UI (same look as the Cowork artifact, but talks to /api endpoints)
# ---------------------------------------------------------------------------

HTML = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>🎯 Jobs Triage</title>
<link rel="icon" type="image/svg+xml" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>%F0%9F%8E%AF</text></svg>">
<style>
:root { color-scheme: light; }
* { box-sizing: border-box; }
html, body { margin: 0; padding: 0; background: #fafbfc; color: #0f172a; font-family: -apple-system, BlinkMacSystemFont, "Inter", "Segoe UI", system-ui, sans-serif; font-size: 14px; }
body { padding: 0 16px; max-width: 920px; margin: 0 auto; }
.top { display: flex; align-items: center; justify-content: space-between; gap: 12px; padding: 18px 4px 12px; border-bottom: 1px solid #e2e8f0; position: sticky; top: 0; background: #fafbfc; z-index: 5; }
.title { font-size: 18px; font-weight: 600; letter-spacing: -0.01em; }
.subtitle { font-size: 12px; color: #64748b; margin-top: 2px; }
.stats { display: flex; gap: 16px; align-items: center; }
.stat { display: flex; flex-direction: column; align-items: flex-end; }
.stat-num { font-size: 20px; font-weight: 600; line-height: 1; font-variant-numeric: tabular-nums; }
.stat-label { font-size: 10px; color: #64748b; text-transform: uppercase; letter-spacing: 0.04em; margin-top: 2px; }
.stat.pending .stat-num { color: #0f172a; }
.stat.yes .stat-num { color: #16a34a; }
.stat.no .stat-num { color: #94a3b8; }
.stat.maybe .stat-num { color: #d97706; }

.filters { display: flex; gap: 6px; padding: 12px 4px; flex-wrap: wrap; }
.chip { padding: 5px 11px; border-radius: 999px; border: 1px solid #e2e8f0; background: white; cursor: pointer; font-size: 12px; font-weight: 500; color: #475569; user-select: none; transition: all 0.15s; }
.chip:hover { border-color: #cbd5e1; background: #f1f5f9; }
.chip.active { background: #0f172a; border-color: #0f172a; color: white; }
.chip .count { opacity: 0.6; margin-left: 4px; font-variant-numeric: tabular-nums; }
.chip.active .count { opacity: 0.85; }

.add-form { background: white; border: 1px solid #e2e8f0; border-radius: 10px; padding: 14px; margin: 4px 0 14px; display: flex; flex-direction: column; gap: 10px; }
.add-row { display: flex; gap: 10px; }
.add-row input { flex: 1; }
.add-form input, .add-form textarea { padding: 8px 12px; border: 1px solid #e2e8f0; border-radius: 7px; font: inherit; color: #0f172a; background: #fafbfc; }
.add-form input:focus, .add-form textarea:focus { outline: none; border-color: #0f172a; background: white; }
.add-form textarea { resize: vertical; min-height: 120px; font-family: SF Mono, Monaco, monospace; font-size: 12px; line-height: 1.5; }
.add-actions { display: flex; align-items: center; gap: 10px; }
.add-hint { font-size: 11px; color: #64748b; flex: 1; }
.add-actions .btn { padding: 7px 14px; }

.cards { display: flex; flex-direction: column; gap: 10px; padding: 4px 0 100px; }
.card { background: white; border: 1px solid #e2e8f0; border-radius: 10px; padding: 14px 16px; transition: opacity 0.25s, transform 0.25s, max-height 0.25s; overflow: hidden; }
.card.removing { opacity: 0; transform: translateX(20px); max-height: 0; padding-top: 0; padding-bottom: 0; margin-top: -10px; border-width: 0; }
.card-head { display: flex; align-items: flex-start; justify-content: space-between; gap: 8px; }
.company { font-size: 16px; font-weight: 600; letter-spacing: -0.01em; line-height: 1.2; }
.role { font-size: 13px; color: #334155; margin-top: 4px; line-height: 1.35; }
.meta { display: flex; flex-wrap: wrap; gap: 4px 12px; margin-top: 8px; font-size: 12px; color: #64748b; }
.meta-item { display: inline-flex; align-items: center; gap: 4px; }
.reasoning { font-size: 12px; color: #475569; margin-top: 8px; font-style: italic; border-left: 2px solid #e2e8f0; padding-left: 8px; }

.tier { font-size: 10px; font-weight: 700; padding: 3px 8px; border-radius: 999px; letter-spacing: 0.04em; text-transform: uppercase; white-space: nowrap; }
.tier-A { background: #dcfce7; color: #15803d; }
.tier-wedge { background: #dbeafe; color: #1d4ed8; }
.tier-unsure { background: #fef3c7; color: #b45309; }

.actions { display: flex; gap: 6px; margin-top: 12px; flex-wrap: wrap; }
.btn { padding: 7px 12px; border-radius: 7px; border: 1px solid #e2e8f0; background: white; color: #0f172a; cursor: pointer; font-size: 12px; font-weight: 500; font-family: inherit; transition: all 0.12s; display: inline-flex; align-items: center; gap: 5px; text-decoration: none; }
.btn:hover { border-color: #cbd5e1; background: #f8fafc; }
.btn-apply { background: #0f172a; color: white; border-color: #0f172a; }
.btn-apply:hover { background: #1e293b; border-color: #1e293b; }
.btn-yes { color: #15803d; border-color: #bbf7d0; }
.btn-yes:hover { background: #f0fdf4; border-color: #86efac; }
.btn-no { color: #b91c1c; border-color: #fecaca; }
.btn-no:hover { background: #fef2f2; border-color: #fca5a5; }
.btn-maybe { color: #b45309; border-color: #fde68a; }
.btn-maybe:hover { background: #fffbeb; border-color: #fcd34d; }
.btn-undo { color: #475569; }
.btn-applied { color: #5b21b6; border-color: #ddd6fe; }
.btn-applied:hover { background: #f5f3ff; border-color: #c4b5fd; }
.applied-stamp { background: #f5f3ff; color: #5b21b6; padding: 2px 8px; border-radius: 4px; font-weight: 500; }
.row-menu-trigger { background: none; border: none; color: #94a3b8; cursor: pointer; font-size: 16px; font-weight: 700; padding: 2px 6px; border-radius: 5px; line-height: 1; font-family: inherit; }
.row-menu-trigger:hover { background: #f1f5f9; color: #0f172a; }
.row-menu { position: absolute; right: 0; top: 28px; background: white; border: 1px solid #e2e8f0; border-radius: 8px; box-shadow: 0 8px 24px rgba(15,23,42,0.12); min-width: 220px; z-index: 50; padding: 4px; display: none; }
.row-menu.open { display: block; }
.row-menu button { display: block; width: 100%; text-align: left; padding: 8px 12px; background: none; border: none; border-radius: 5px; font: inherit; font-size: 12px; color: #334155; cursor: pointer; }
.row-menu button:hover { background: #f8fafc; }
.row-menu button.danger { color: #b91c1c; }
.row-menu button.danger:hover { background: #fef2f2; }
.btn-verify-date { color: #1d4ed8; border-color: #dbeafe; padding: 2px 7px; font-size: 11px; margin-left: 4px; line-height: 1.3; }
.btn-verify-date:hover:not(:disabled) { background: #eff6ff; border-color: #93c5fd; }
.date-verifying { color: #1d4ed8; display: inline-flex; align-items: center; gap: 5px; }
.date-verifying .tail-spin { width: 11px; height: 11px; border: 2px solid #93c5fd; border-top-color: transparent; border-radius: 50%; animation: spin 0.6s linear infinite; }
.date-verified { color: #15803d; font-weight: 500; }
.date-unknown { color: #b45309; }
.btn-outreach { color: #be185d; border-color: #fbcfe8; }
.btn-outreach:hover:not(:disabled) { background: #fdf2f8; border-color: #f9a8d4; }
.btn-outreach:disabled { opacity: 0.7; cursor: default; }
.btn-outreach-ready { color: #0f766e; border-color: #99f6e4; }
.btn-outreach-ready:hover { background: #f0fdfa; border-color: #5eead4; }
.btn-outreach-sent { color: #15803d; border-color: #86efac; background: #f0fdf4; }

/* Outreach modal */
.modal-bg { position: fixed; inset: 0; background: rgba(15,23,42,0.5); z-index: 200; display: flex; align-items: flex-start; justify-content: center; padding: 32px 16px; overflow-y: auto; }
.modal-bg.hidden { display: none; }
.modal { background: white; border-radius: 12px; max-width: 1040px; width: 100%; box-shadow: 0 20px 60px rgba(0,0,0,0.3); display: flex; flex-direction: column; max-height: calc(100vh - 64px); }
.modal-head { padding: 16px 20px; border-bottom: 1px solid #e2e8f0; display: flex; align-items: center; justify-content: space-between; gap: 12px; }
.modal-title { font-size: 16px; font-weight: 600; }
.modal-subtitle { font-size: 12px; color: #64748b; margin-top: 2px; }
.modal-close { background: none; border: none; font-size: 22px; cursor: pointer; color: #94a3b8; padding: 4px 10px; border-radius: 6px; }
.modal-close:hover { background: #f1f5f9; color: #0f172a; }
.modal-body { padding: 16px 20px; overflow-y: auto; flex: 1; }
.modal-foot { padding: 12px 20px; border-top: 1px solid #e2e8f0; display: flex; align-items: center; justify-content: space-between; gap: 10px; background: #f8fafc; border-radius: 0 0 12px 12px; }

.tmpl-block { display: flex; flex-direction: column; gap: 6px; margin-bottom: 14px; padding: 10px 12px; background: #f8fafc; border-radius: 8px; border: 1px solid #e2e8f0; }
.tmpl-head { font-size: 11px; font-weight: 600; color: #475569; letter-spacing: 0.04em; text-transform: uppercase; display: flex; justify-content: space-between; }
.tmpl-block textarea { width: 100%; padding: 6px 8px; border: 1px solid #e2e8f0; border-radius: 6px; font: inherit; font-size: 12px; line-height: 1.4; min-height: 60px; resize: vertical; background: white; }
.tmpl-block textarea:focus { outline: none; border-color: #0f172a; }
.tmpl-hint { font-size: 10px; color: #94a3b8; }
.char-counter { font-variant-numeric: tabular-nums; font-size: 10px; color: #64748b; }
.char-counter.over { color: #b91c1c; font-weight: 600; }

.brief-row { display: flex; align-items: center; gap: 10px; margin-bottom: 14px; padding: 10px 12px; background: #fef3c7; border-radius: 8px; }
.brief-row label { font-size: 12px; font-weight: 600; color: #92400e; white-space: nowrap; }
.brief-row input { flex: 1; padding: 6px 10px; border: 1px solid #fde68a; border-radius: 6px; font: inherit; font-size: 13px; background: white; }
.brief-row input:focus { outline: none; border-color: #b45309; }

.lead-row { display: grid; grid-template-columns: 28px 1fr 1fr 110px 80px; gap: 10px; align-items: start; padding: 12px; border-bottom: 1px solid #f1f5f9; }
.lead-row:last-child { border-bottom: none; }
.lead-row.unapproved { opacity: 0.55; background: #fafafa; }
.lead-row.unapproved .lead-msg textarea { background: #f8fafc; }
.lead-cb { width: 18px; height: 18px; margin-top: 4px; cursor: pointer; }
.lead-info .lead-name { font-weight: 600; font-size: 13px; }
.lead-info .lead-title { font-size: 11px; color: #64748b; margin-top: 2px; }
.lead-info .lead-summary { font-size: 11px; color: #475569; margin-top: 4px; line-height: 1.45; }
.lead-info .lead-tags { margin-top: 6px; display: flex; gap: 4px; flex-wrap: wrap; }
.lead-tag { font-size: 9px; font-weight: 600; letter-spacing: 0.05em; text-transform: uppercase; padding: 2px 6px; border-radius: 4px; }
.tag-1st { background: #dbeafe; color: #1d4ed8; }
.tag-2nd { background: #fef3c7; color: #b45309; }
.tag-3rd { background: #fee2e2; color: #b91c1c; }
.tag-recruiter { background: #dcfce7; color: #15803d; }
.tag-founder { background: #fce7f3; color: #be185d; }
.tag-engineer { background: #ede9fe; color: #5b21b6; }
.lead-msg textarea { width: 100%; min-height: 60px; font-family: SF Mono, Monaco, monospace; font-size: 11px; padding: 6px 8px; border: 1px solid #e2e8f0; border-radius: 6px; resize: vertical; line-height: 1.5; background: white; }
.lead-status { font-size: 11px; font-weight: 600; padding: 3px 8px; border-radius: 6px; display: inline-block; text-align: center; }
.status-pending-send { background: #fef3c7; color: #92400e; }
.status-sent { background: #dcfce7; color: #15803d; }
.status-failed { background: #fee2e2; color: #b91c1c; }
.status-skipped { background: #f1f5f9; color: #64748b; }
.status-limit-reached { background: #fef2f2; color: #b91c1c; }
.lead-link { font-size: 11px; color: #1d4ed8; text-decoration: none; padding: 4px 8px; border: 1px solid #dbeafe; border-radius: 6px; display: inline-block; text-align: center; }
.lead-link:hover { background: #dbeafe; }

.modal-empty { padding: 60px 20px; text-align: center; color: #64748b; font-size: 13px; }

.lead-row-head { display: grid; grid-template-columns: 28px 1fr 1fr 110px 80px; gap: 10px; padding: 8px 12px; font-size: 10px; font-weight: 600; letter-spacing: 0.04em; text-transform: uppercase; color: #94a3b8; border-bottom: 1px solid #e2e8f0; }

.decision-tag { font-size: 11px; font-weight: 500; padding: 2px 8px; border-radius: 4px; }
.decision-yes { background: #dcfce7; color: #15803d; }
.decision-no { background: #fee2e2; color: #b91c1c; }
.decision-maybe { background: #fef3c7; color: #b45309; }

.tailor-state { margin-top: 10px; padding: 10px 12px; border-radius: 8px; display: flex; align-items: center; gap: 10px; font-size: 12px; }
.tailor-pending { background: #fef9c3; color: #854d0e; border: 1px solid #fde68a; }
.tailor-ready { background: #ecfdf5; color: #047857; border: 1px solid #a7f3d0; }
.tailor-error { background: #fef2f2; color: #b91c1c; border: 1px solid #fecaca; }
.tailor-state .tail-spin { width: 13px; height: 13px; border: 2px solid #d6b441; border-top-color: transparent; border-radius: 50%; animation: spin 0.6s linear infinite; }
.tailor-state .open-link { margin-left: auto; padding: 4px 10px; background: #047857; color: white; text-decoration: none; border-radius: 6px; font-weight: 500; font-size: 12px; }
.tailor-state .open-link:hover { background: #065f46; }

.state { text-align: center; padding: 60px 20px; color: #64748b; }
.state-emoji { font-size: 36px; margin-bottom: 12px; }
.state-title { font-size: 15px; font-weight: 600; color: #0f172a; margin-bottom: 6px; }
.state-text { font-size: 13px; max-width: 360px; margin: 0 auto; line-height: 1.5; }
.retry-btn { margin-top: 14px; padding: 8px 16px; border-radius: 7px; border: 1px solid #0f172a; background: #0f172a; color: white; cursor: pointer; font-size: 13px; font-weight: 500; font-family: inherit; }

.toast { position: fixed; bottom: 20px; left: 50%; transform: translateX(-50%); background: #0f172a; color: white; padding: 10px 18px; border-radius: 8px; font-size: 13px; font-weight: 500; box-shadow: 0 8px 24px rgba(15, 23, 42, 0.2); z-index: 100; display: flex; gap: 12px; align-items: center; transition: opacity 0.2s, transform 0.2s; }
.toast.error { background: #b91c1c; }
.toast.hidden { opacity: 0; pointer-events: none; transform: translateX(-50%) translateY(8px); }
.toast .toast-undo { background: rgba(255,255,255,0.18); color: white; border: none; padding: 4px 10px; border-radius: 5px; cursor: pointer; font-size: 12px; font-weight: 500; }
.toast .toast-undo:hover { background: rgba(255,255,255,0.28); }

.spinner { display: inline-block; width: 14px; height: 14px; border: 2px solid #e2e8f0; border-top-color: #0f172a; border-radius: 50%; animation: spin 0.6s linear infinite; vertical-align: middle; margin-right: 6px; }
@keyframes spin { to { transform: rotate(360deg); } }
.loading-block { display: flex; gap: 10px; align-items: center; justify-content: center; padding: 30px; color: #64748b; font-size: 13px; }
.reload-btn { padding: 5px 10px; border-radius: 6px; border: 1px solid #e2e8f0; background: white; color: #475569; cursor: pointer; font-size: 12px; font-weight: 500; font-family: inherit; }
.reload-btn:hover { background: #f1f5f9; }
a { color: inherit; }
</style>
</head>
<body>

<div class="top">
  <div>
    <div class="title">Jobs Triage</div>
    <div class="subtitle" id="subtitle">Loading…</div>
  </div>
  <div class="stats">
    <div class="stat pending"><div class="stat-num" id="count-new">·</div><div class="stat-label">New</div></div>
    <div class="stat yes"><div class="stat-num" id="count-applied">·</div><div class="stat-label">Applied</div></div>
    <div class="stat yes"><div class="stat-num" id="count-outreached">·</div><div class="stat-label">Outreached</div></div>
    <div class="stat no"><div class="stat-num" id="count-passed">·</div><div class="stat-label">Passed</div></div>
    <a href="/runs" class="reload-btn" style="text-decoration:none;color:inherit;padding:5px 10px;" title="Live viewer — watch any Claude run (sourcing, tailor, outreach, verify) as it happens">📡 Live</a>
    <button class="reload-btn" id="kill-all-btn" title="Close all stuck Claude terminals (Terminal.app only, not iTerm2)">🧹</button>
    <button class="reload-btn" id="reload-btn" title="Reload">↻</button>
  </div>
</div>

<div class="filters" id="filters">
  <button class="chip active" data-filter="new">🆕 New<span class="count" id="ch-new">·</span></button>
  <button class="chip" data-filter="maybe">○ Maybe<span class="count" id="ch-maybe">·</span></button>
  <button class="chip" data-filter="queued">✓ Queued<span class="count" id="ch-queued">·</span></button>
  <button class="chip" data-filter="applied">✈ Applied<span class="count" id="ch-applied">·</span></button>
  <button class="chip" data-filter="reaching">🎯 Reaching out<span class="count" id="ch-reaching">·</span></button>
  <button class="chip" data-filter="outreached">📬 Outreached<span class="count" id="ch-outreached">·</span></button>
  <button class="chip" data-filter="passed">✗ Passed<span class="count" id="ch-passed">·</span></button>
  <button class="chip" data-filter="all">All<span class="count" id="ch-all">·</span></button>
  <button class="chip" id="add-toggle" style="margin-left:auto;background:#0f172a;color:white;border-color:#0f172a;">+ Add JD</button>
</div>

<div class="add-form" id="add-form" style="display:none">
  <input type="text" id="add-jd-input" placeholder="Paste a job URL (e.g. boards.greenhouse.io/...) or the full JD text — Claude figures out the rest." autocomplete="off" autofocus>
  <div class="add-actions">
    <span class="add-hint">Saves a row dated today, decision=yes, then spawns Claude. If you paste a URL, Claude fetches the JD; if you paste JD text, it uses that directly. Company / role / location all auto-extracted.</span>
    <button class="btn" id="add-cancel">Cancel</button>
    <button class="btn btn-apply" id="add-submit">Add & Start Tailoring</button>
  </div>
</div>

<div id="content">
  <div class="loading-block"><span class="spinner"></span>Loading jobs…</div>
</div>

<div class="toast hidden" id="toast"><span id="toast-msg"></span><button class="toast-undo" id="toast-undo" style="display:none">Undo</button></div>

<!-- Outreach: Find Leads count picker -->
<div class="modal-bg hidden" id="find-modal">
  <div class="modal" style="max-width:440px">
    <div class="modal-head">
      <div>
        <div class="modal-title">Find LinkedIn leads</div>
        <div class="modal-subtitle" id="find-modal-sub">…</div>
      </div>
      <button class="modal-close" data-close="find-modal">×</button>
    </div>
    <div class="modal-body">
      <div style="font-size:13px;color:#475569;margin-bottom:14px">How many recruiters / founders / engineers should Claude find? Claude prioritizes recruiters (1st-degree connections first), then falls back to founders for startups, then engineers.</div>
      <div style="display:flex;gap:8px;margin-bottom:12px">
        <button class="btn" data-find-count="5">5 leads</button>
        <button class="btn" data-find-count="10">10 leads</button>
        <button class="btn" data-find-count="15">15 leads</button>
        <button class="btn" data-find-count="20">20 leads</button>
      </div>
      <div style="font-size:11px;color:#94a3b8">Claude will open Terminal + Chrome MCP and run for a few minutes. You'll see results when stage = leads-ready.</div>
    </div>
  </div>
</div>

<!-- Outreach: View Leads / send -->
<div class="modal-bg hidden" id="view-modal">
  <div class="modal">
    <div class="modal-head">
      <div>
        <div class="modal-title" id="view-modal-title">Leads</div>
        <div class="modal-subtitle" id="view-modal-sub">…</div>
      </div>
      <button class="modal-close" data-close="view-modal">×</button>
    </div>
    <div class="modal-body" id="view-modal-body">
      <div class="loading-block"><span class="spinner"></span>Loading leads…</div>
    </div>
    <div class="modal-foot">
      <div style="display:flex;gap:8px">
        <button class="btn" id="view-find-more">+ Find more leads</button>
        <button class="btn btn-yes" id="view-approve-all">✓ Approve all</button>
        <button class="btn btn-no" id="view-unapprove-all">✗ Unapprove all</button>
      </div>
      <div style="display:flex;gap:8px;align-items:center">
        <span id="view-approve-summary" style="font-size:12px;color:#475569"></span>
        <button class="btn" id="view-save">💾 Save</button>
        <button class="btn btn-apply" id="view-send">📨 Send approved</button>
      </div>
    </div>
  </div>
</div>

<script>
(function() {
  let rows = [];
  // Migrate any old filter names from localStorage so existing users don't land on a dead chip.
  const FILTER_MIGRATE = { pending: 'new', yes: 'queued', no: 'passed', A: 'new', wedge: 'new', unsure: 'new' };
  let activeFilter = localStorage.getItem('jobs-triage-filter') || 'new';
  if (FILTER_MIGRATE[activeFilter]) activeFilter = FILTER_MIGRATE[activeFilter];
  let lastUndo = null;
  const $ = (id) => document.getElementById(id);
  const content = $('content');

  async function fetchJobs() {
    const res = await fetch('/api/jobs', { cache: 'no-store' });
    if (!res.ok) throw new Error('HTTP ' + res.status);
    return await res.json();
  }
  async function postDecision(id, decision) {
    const res = await fetch('/api/decide', {
      method: 'POST',
      headers: { 'Content-Type': 'application/json' },
      body: JSON.stringify({ id: Number(id), decision }),
    });
    const data = await res.json().catch(() => ({}));
    if (!res.ok) throw new Error(data.error || ('HTTP ' + res.status));
    return data;
  }

  function escapeHtml(s) {
    return (s == null ? '' : String(s)).replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":"&#39;"}[c]));
  }
  function tierClass(t) {
    const lo = (t || '').toLowerCase();
    if (lo === 'a') return 'tier-A';
    if (lo === 'wedge') return 'tier-wedge';
    if (lo === 'unsure') return 'tier-unsure';
    return '';
  }
  function tierLabel(t) {
    const lo = (t || '').toLowerCase();
    if (lo === 'a') return 'Tier A';
    if (lo === 'wedge') return 'Wedge';
    if (lo === 'unsure') return 'Unsure';
    return t || '';
  }
  function decisionTag(d) {
    if (d === 'yes') return '<span class="decision-tag decision-yes">✓ Yes</span>';
    if (d === 'no') return '<span class="decision-tag decision-no">✗ No</span>';
    if (d === 'maybe') return '<span class="decision-tag decision-maybe">○ Maybe</span>';
    return '';
  }

  // Pipeline stage classifier — returns one of:
  //   'new'         → undecided (decision=pending)
  //   'maybe'       → deferred (decision=maybe)
  //   'queued'      → approved but not applied yet
  //   'applied'     → applied, no outreach activity
  //   'reaching'    → outreach in flight (finding leads / sending messages)
  //   'outreached'  → outreach complete (sent)
  //   'passed'      → rejected (decision=no)
  function pipelineStage(r) {
    const d = (r.decision || 'pending').toLowerCase();
    if (d === 'pending') return 'new';
    if (d === 'maybe') return 'maybe';
    if (d === 'no') return 'passed';
    // decision === 'yes' from here on
    if (!r.applied_date) return 'queued';
    const stage = (r.outreach_stage || '').toLowerCase();
    if (stage === 'finding' || stage === 'leads-ready' || stage === 'sending') return 'reaching';
    if (stage === 'sent') return 'outreached';
    return 'applied';
  }

  function computeCounts() {
    const counts = { new: 0, maybe: 0, queued: 0, applied: 0, reaching: 0, outreached: 0, passed: 0, all: rows.length };
    for (const r of rows) {
      const s = pipelineStage(r);
      counts[s] = (counts[s] || 0) + 1;
    }
    return counts;
  }

  function applyFilter(r, filter) {
    if (filter === 'all') return true;
    return pipelineStage(r) === filter;
  }

  function tailorStateHtml(r) {
    const rv = (r.resume_version || '').trim();
    if (!rv) return '';
    if (rv === 'tailoring') {
      return `<div class="tailor-state tailor-pending"><span class="tail-spin"></span><span>Tailoring resume… Claude Code is iterating on the 1-pager. Refreshes automatically.</span><button class="btn btn-no" data-act="cancel-tailor" data-id="${escapeHtml(r.id)}" style="margin-left:auto;">Cancel</button></div>`;
    }
    if (rv.toLowerCase().startsWith('error:')) {
      return `<div class="tailor-state tailor-error"><span>⚠️ ${escapeHtml(rv)}</span><button class="btn btn-yes" data-act="retry-tailor" data-id="${escapeHtml(r.id)}" style="margin-left:auto;">Retry</button></div>`;
    }
    // Path to a folder or .pdf. Build a URL-safe link to the PDF.
    let rel = rv;
    // Strip whatever absolute prefix is in the xlsx cell. The /resume-files/ HTTP
    // route serves from RESUME_ROOT, so we want everything from "versions/" onward
    // regardless of where the buyer's repo lives on disk.
    const vIdx = rel.indexOf('/versions/');
    if (vIdx >= 0) rel = rel.slice(vIdx + 1);
    // URL-encode each path segment (filenames have spaces in the new convention)
    function encodePath(p) { return p.split('/').map(s => encodeURIComponent(s)).join('/'); }
    let pdfUrl;
    if (rel.endsWith('.pdf')) {
      pdfUrl = '/resume-files/' + encodePath(rel);
    } else {
      // legacy: folder path without filename — fall back to tailored.pdf
      if (!rel.endsWith('/')) rel = rel + '/';
      pdfUrl = '/resume-files/' + encodePath(rel) + 'tailored.pdf';
    }
    return `<div class="tailor-state tailor-ready"><span>✓ Resume ready</span><a class="open-link" target="_blank" href="${escapeHtml(pdfUrl)}">Open PDF ↗</a></div>`;
  }

  function cardHtml(r) {
    const dec = (r.decision || 'pending').toLowerCase();
    const isPending = dec === 'pending';
    const tier = tierClass(r.tier);
    const tlabel = tierLabel(r.tier);
    return `
      <div class="card" data-id="${escapeHtml(r.id)}">
        <div class="card-head">
          <div style="min-width: 0; flex: 1;">
            <div class="company">${escapeHtml(r.company)}</div>
            <div class="role">${escapeHtml(r.role)}</div>
          </div>
          <div style="display:flex; flex-direction:column; gap:6px; align-items:flex-end; position:relative;">
            <div style="display:flex; gap:6px; align-items:center;">
              ${tier ? `<span class="tier ${tier}">${escapeHtml(tlabel)}</span>` : ''}
              <button class="row-menu-trigger" data-act="row-menu" data-id="${escapeHtml(r.id)}" title="More actions">⋯</button>
            </div>
            ${!isPending ? decisionTag(dec) : ''}
          </div>
        </div>
        <div class="meta">
          ${r.location ? `<span class="meta-item">📍 ${escapeHtml(r.location)}</span>` : ''}
          ${r.comp && r.comp !== 'unknown' ? `<span class="meta-item">💰 ${escapeHtml(r.comp)}</span>` : ''}
          ${r.source ? `<span class="meta-item">🔗 ${escapeHtml(r.source)}</span>` : ''}
          ${dateMetaHtml(r)}
          ${r.date_sourced && r.date_sourced !== r.posted_date ? `<span class="meta-item" title="Saved on">📥 ${escapeHtml(r.date_sourced)}</span>` : ''}
          ${r.applied_date ? `<span class="meta-item applied-stamp" title="Applied on">✈ Applied ${escapeHtml(r.applied_date)}</span>` : ''}
        </div>
        ${r.reasoning ? `<div class="reasoning">${escapeHtml(r.reasoning)}</div>` : ''}
        ${tailorStateHtml(r)}
        <div class="actions">
          ${r.apply_url ? `<a class="btn btn-apply" target="_blank" rel="noopener" href="${escapeHtml(r.apply_url)}">Apply ↗</a>` : ''}
          ${isPending ? `
            <button class="btn btn-yes" data-act="yes" data-id="${escapeHtml(r.id)}">✓ Yes</button>
            <button class="btn btn-maybe" data-act="maybe" data-id="${escapeHtml(r.id)}">○ Maybe</button>
            <button class="btn btn-no" data-act="no" data-id="${escapeHtml(r.id)}">✗ No</button>
          ` : `
            <button class="btn btn-undo" data-act="pending" data-id="${escapeHtml(r.id)}">↺ Reset to pending</button>
          `}
          ${dec !== 'no' ? (
            r.applied_date
              ? `<button class="btn btn-undo" data-act="unapply" data-id="${escapeHtml(r.id)}" title="Clear applied date">↺ Unmark applied</button>`
              : `<button class="btn btn-applied" data-act="apply-now" data-id="${escapeHtml(r.id)}">✈ Mark applied</button>`
          ) : ''}
          ${r.applied_date ? outreachButtonHtml(r) : ''}
        </div>
      </div>
    `;
  }

  function dateMetaHtml(r) {
    const aggDate = r.posted_date || '';
    const verified = (r.posted_date_verified || '').trim();
    if (!aggDate && !verified) return '';
    if (verified === 'verifying') {
      return `<span class="meta-item date-verifying" title="Verifying actual posted date..."><span class="tail-spin"></span>Verifying date…</span>`;
    }
    if (verified === 'unknown') {
      return `<span class="meta-item date-unknown" title="Could not find a posted date on the source ATS page">⚠ Posted date unknown${aggDate ? ` (aggregator said ${escapeHtml(aggDate)})` : ''}</span><button class="btn btn-verify-date" data-act="verify-date" data-id="${escapeHtml(r.id)}" title="Re-verify">🔄</button>`;
    }
    if (verified && /^\d{4}-\d{2}-\d{2}$/.test(verified)) {
      const mismatch = aggDate && verified !== aggDate;
      return `<span class="meta-item date-verified" title="${mismatch ? 'Verified from source. Aggregator reported a different date: ' + aggDate : 'Verified from source ATS page'}">✓ Posted ${escapeHtml(verified)}${mismatch ? ` <span style="color:#94a3b8">(agg said ${escapeHtml(aggDate)})</span>` : ''}</span>`;
    }
    // Not yet verified — show aggregator date with a verify button
    return `<span class="meta-item" title="From aggregator — click 🔍 to verify from the source">📅 ${escapeHtml(aggDate)}</span><button class="btn btn-verify-date" data-act="verify-date" data-id="${escapeHtml(r.id)}" title="Verify actual posted date from source ATS">🔍 Verify</button>`;
  }

  function outreachButtonHtml(r) {
    const stage = (r.outreach_stage || '').toLowerCase();
    const total = parseInt(r.outreach_lead_count || 0, 10);
    const sent = parseInt(r.outreach_sent || 0, 10);
    const approved = parseInt(r.outreach_approved || 0, 10);
    if (!stage) {
      return `<button class="btn btn-outreach" data-act="outreach-find" data-id="${escapeHtml(r.id)}" title="Find LinkedIn recruiters/founders/engineers">🎯 Outreach</button>`;
    }
    if (stage === 'finding') {
      return `<button class="btn btn-outreach" disabled title="Claude is searching LinkedIn"><span class="tail-spin" style="margin-right:6px"></span>Finding leads…</button>`;
    }
    if (stage === 'sending') {
      return `<button class="btn btn-outreach" disabled title="Sending messages"><span class="tail-spin" style="margin-right:6px"></span>Sending… (${sent}/${approved})</button>`;
    }
    if (stage === 'sent') {
      return `<button class="btn btn-outreach btn-outreach-sent" data-act="outreach-view" data-id="${escapeHtml(r.id)}" title="View sent outreach">✓ Outreach sent (${sent})</button>`;
    }
    // leads-ready (or anything else with leads): show View Leads
    return `<button class="btn btn-outreach btn-outreach-ready" data-act="outreach-view" data-id="${escapeHtml(r.id)}" title="${total} leads found, ${approved} approved">👁 View leads (${total})</button>`;
  }

  function render() {
    const counts = computeCounts();
    // Top stats (funnel snapshot)
    $('count-new').textContent = counts.new;
    // Applied total = everything from "applied" state onward (applied + reaching + outreached)
    $('count-applied').textContent = counts.applied + counts.reaching + counts.outreached;
    $('count-outreached').textContent = counts.outreached;
    $('count-passed').textContent = counts.passed;
    // Per-chip counts
    $('ch-new').textContent = counts.new;
    $('ch-maybe').textContent = counts.maybe;
    $('ch-queued').textContent = counts.queued;
    $('ch-applied').textContent = counts.applied;
    $('ch-reaching').textContent = counts.reaching;
    $('ch-outreached').textContent = counts.outreached;
    $('ch-passed').textContent = counts.passed;
    $('ch-all').textContent = counts.all;
    const dt = new Date();
    $('subtitle').textContent = `${rows.length} total · loaded ${dt.toLocaleTimeString([], {hour:'2-digit', minute:'2-digit'})}`;

    document.querySelectorAll('.chip').forEach(c => {
      c.classList.toggle('active', c.dataset.filter === activeFilter);
    });

    const filtered = rows.filter(r => applyFilter(r, activeFilter));
    if (!filtered.length) {
      const emptyMsgs = {
        new:        { e: '🎯', t: 'All caught up',     m: 'Nothing left to triage. The noon-PT scheduled task will surface new finds.' },
        queued:     { e: '📋', t: 'Nothing queued',    m: 'No approved jobs waiting to be applied to.' },
        applied:    { e: '✈',  t: 'No active apps',    m: 'Jobs you mark as applied (but haven\'t outreached yet) show up here.' },
        reaching:   { e: '🎯', t: 'No outreach yet',   m: 'Jobs with outreach in flight (finding leads / sending) show up here.' },
        outreached: { e: '📬', t: 'No completed outreach', m: 'Jobs with outreach fully sent show up here.' },
        maybe:      { e: '○',  t: 'No deferred jobs',  m: 'Jobs you marked Maybe show up here.' },
        passed:     { e: '✗',  t: 'No passed jobs',    m: 'Jobs you marked No show up here.' },
        all:        { e: '∅',  t: 'No jobs at all',    m: 'The sheet is empty. Trigger a sourcing run.' },
      };
      const msg = emptyMsgs[activeFilter] || { e: '∅', t: 'Nothing matches this filter', m: 'Pick another filter or wait for the next run.' };
      content.innerHTML = `<div class="state"><div class="state-emoji">${msg.e}</div><div class="state-title">${msg.t}</div><div class="state-text">${msg.m}</div></div>`;
      return;
    }
    content.innerHTML = `<div class="cards">${filtered.map(cardHtml).join('')}</div>`;
  }

  let pollTimer = null;
  function maybeStartPolling() {
    const anyTailoring = rows.some(r => (r.resume_version || '').trim() === 'tailoring');
    const anyOutreachActive = rows.some(r => {
      const s = (r.outreach_stage || '').toLowerCase();
      return s === 'finding' || s === 'sending';
    });
    const anyVerifying = rows.some(r => (r.posted_date_verified || '').trim() === 'verifying');
    if ((anyTailoring || anyOutreachActive || anyVerifying) && !pollTimer) {
      pollTimer = setInterval(() => loadJobs(true), 8000);
    } else if (!anyTailoring && !anyOutreachActive && !anyVerifying && pollTimer) {
      clearInterval(pollTimer); pollTimer = null;
    }
  }

  async function loadJobs(silent) {
    if (!silent) content.innerHTML = `<div class="loading-block"><span class="spinner"></span>Loading jobs…</div>`;
    try {
      const data = await fetchJobs();
      if (data.error) throw new Error(data.error);
      rows = data.rows || [];
      // Sort: by most recent posted date (verified > aggregator > date_sourced > applied_date),
      // newest first. Tiebreak by id desc. Always-by-date means the freshest opportunities
      // surface to the top no matter which pipeline filter the user picks.
      function rowDate(r) {
        return r.posted_date_verified && /^\d{4}-\d{2}-\d{2}$/.test(r.posted_date_verified)
          ? r.posted_date_verified
          : (r.posted_date || r.date_sourced || r.applied_date || '');
      }
      rows.sort((a, b) => {
        const da = rowDate(a);
        const db = rowDate(b);
        if (db !== da) return db.localeCompare(da);  // YYYY-MM-DD strings compare correctly lexically
        return parseInt(b.id) - parseInt(a.id);
      });
      render();
      maybeStartPolling();
    } catch (e) {
      content.innerHTML = `<div class="state"><div class="state-emoji">⚠️</div><div class="state-title">Couldn't load jobs</div><div class="state-text">${escapeHtml(e.message)}</div><button class="retry-btn" id="retry-btn">Retry</button></div>`;
      $('subtitle').textContent = 'Error';
      const r = document.getElementById('retry-btn');
      if (r) r.addEventListener('click', () => loadJobs());
    }
  }

  async function decide(id, decision) {
    const row = rows.find(r => String(r.id) === String(id));
    if (!row) return;
    if (decision === 'cancel-tailor') {
      await cancelTailor(id);
      return;
    }
    if (decision === 'apply-now') {
      await markApplied(id, null);
      return;
    }
    if (decision === 'unapply') {
      await markApplied(id, '');
      return;
    }
    if (decision === 'retry-tailor') {
      // Reset resume_version and re-trigger Yes
      row.resume_version = '';
      decision = 'yes';
    }
    if (decision === 'verify-date') {
      await triggerVerifyDate(id);
      return;
    }
    const prev = row.decision || 'pending';
    const prevRv = row.resume_version || '';
    row.decision = decision;
    // optimistic: if we're saying yes, show the tailoring state immediately
    if (decision === 'yes' && (!(row.resume_version || '').startsWith('/'))) {
      row.resume_version = 'tailoring';
    }
    const cardEl = document.querySelector(`.card[data-id="${id}"]`);
    const willHide = !applyFilter(row, activeFilter);
    if (cardEl && willHide) cardEl.classList.add('removing');
    try {
      const resp = await postDecision(id, decision);
      lastUndo = { id, prev };
      let msg = `#${id} → ${decision}`;
      if (resp && resp.tailoring_started) msg += ' · tailoring kicked off';
      showToast(msg, true);
      setTimeout(() => { render(); maybeStartPolling(); }, willHide ? 250 : 0);
    } catch (e) {
      // Roll back optimistic UI (decision + the tailoring state we set)
      row.decision = prev;
      row.resume_version = prevRv;
      if (cardEl) cardEl.classList.remove('removing');
      render();
      showToast('Save failed: ' + e.message, false, false);
    }
  }

  async function markApplied(id, applied_date) {
    const row = rows.find(r => String(r.id) === String(id));
    if (!row) return;
    const prev = row.applied_date || '';
    // optimistic update
    row.applied_date = applied_date === null
      ? new Date().toISOString().slice(0, 10)
      : applied_date;
    render();
    try {
      const res = await fetch('/api/applied', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ id: Number(id), applied_date }),
      });
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.error || ('HTTP ' + res.status));
      row.applied_date = data.applied_date || '';
      render();
      showToast(
        row.applied_date
          ? `#${id} marked applied · ${row.applied_date}`
          : `#${id} applied cleared`,
        false
      );
    } catch (e) {
      row.applied_date = prev;
      render();
      showToast('Failed: ' + e.message, false, false);
    }
  }

  async function killTerminals(kind, id) {
    const body = kind === 'job' ? { kind: 'job', id: Number(id) } : { kind: 'all' };
    const label = kind === 'job' ? `for job #${id}` : 'all stuck Claude terminals';
    if (!confirm(`Close ${label}? Any in-flight tasks (tailoring, outreach, verifying) will be cancelled and revert to a usable state. iTerm2 windows are not affected.`)) return;
    try {
      const res = await fetch('/api/kill-terminals', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(body),
      });
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.error || ('HTTP ' + res.status));
      const reset = data.reset || {};
      const resetSummary = Object.entries(reset).filter(([k,v]) => v > 0).map(([k,v]) => `${k}:${v}`).join(', ');
      showToast(`Closed ${data.closed} window(s) · killed ${data.killed} proc(s)${resetSummary ? ' · reset ' + resetSummary : ''}`, false);
      closeAllRowMenus();
      await loadJobs(true);
    } catch (e) {
      showToast('Kill failed: ' + e.message, false, false);
    }
  }

  function closeAllRowMenus() {
    document.querySelectorAll('.row-menu.open').forEach(m => m.classList.remove('open'));
  }

  function toggleRowMenu(id, anchor) {
    closeAllRowMenus();
    const card = anchor.closest('.card');
    if (!card) return;
    // Build the menu inline so it can position relative to the card header
    let menu = card.querySelector('.row-menu');
    if (!menu) {
      menu = document.createElement('div');
      menu.className = 'row-menu';
      menu.innerHTML = `
        <button data-row-act="kill" data-id="${escapeHtml(id)}" class="danger">🧹 Kill terminals for this job</button>
      `;
      anchor.parentElement.appendChild(menu);
    }
    menu.classList.add('open');
  }

  async function triggerVerifyDate(id) {
    const row = rows.find(r => String(r.id) === String(id));
    if (!row) return;
    const prev = row.posted_date_verified || '';
    row.posted_date_verified = 'verifying';
    render();
    try {
      const res = await fetch('/api/verify-date', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ id: Number(id) }),
      });
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.error || ('HTTP ' + res.status));
      showToast(`Verifying #${id}… Claude is opening the source page`, false);
      maybeStartPolling();
    } catch (e) {
      row.posted_date_verified = prev;
      render();
      showToast('Verify failed: ' + e.message, false, false);
    }
  }

  async function cancelTailor(id) {
    try {
      const res = await fetch('/api/cancel', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ id: Number(id) }),
      });
      const data = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(data.error || ('HTTP ' + res.status));
      showToast(`#${id} cancelled · back to pending (close the Terminal window manually)`, false);
      await loadJobs(true);
    } catch (e) {
      showToast('Cancel failed: ' + e.message, false, false);
    }
  }

  async function doUndo() {
    if (!lastUndo) return;
    const { id, prev } = lastUndo;
    lastUndo = null;
    await decide(id, prev);
  }

  let toastTimer;
  function showToast(msg, withUndo, autoHide = true) {
    const t = $('toast');
    $('toast-msg').textContent = msg;
    $('toast-undo').style.display = withUndo ? '' : 'none';
    t.classList.toggle('error', !withUndo && !autoHide);
    t.classList.remove('hidden');
    clearTimeout(toastTimer);
    if (autoHide) toastTimer = setTimeout(() => t.classList.add('hidden'), 3500);
  }

  // ---------- Outreach modals ----------
  let findModalJobId = null;
  let viewModalState = null;  // { jobId, sidecar, dirty }

  function openFindModal(jobId) {
    findModalJobId = jobId;
    const row = rows.find(r => String(r.id) === String(jobId));
    $('find-modal-sub').textContent = row ? `${row.company} — ${row.role}` : '';
    $('find-modal').classList.remove('hidden');
  }
  function closeModal(id) {
    $(id).classList.add('hidden');
    if (id === 'view-modal') viewModalState = null;
    if (id === 'find-modal') findModalJobId = null;
  }

  async function triggerFind(jobId, count, append) {
    try {
      const res = await fetch('/api/outreach-find', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ id: Number(jobId), count, append: !!append }),
      });
      const json = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(json.error || ('HTTP ' + res.status));
      showToast(`Searching LinkedIn for ${count} leads…`, false);
      closeModal('find-modal');
      closeModal('view-modal');
      await loadJobs(true);
      maybeStartPolling();
    } catch (e) {
      showToast('Find failed: ' + e.message, false, false);
    }
  }

  async function openViewModal(jobId) {
    viewModalState = { jobId: Number(jobId), sidecar: null, dirty: false };
    $('view-modal-body').innerHTML = `<div class="loading-block"><span class="spinner"></span>Loading leads…</div>`;
    $('view-modal').classList.remove('hidden');
    try {
      const res = await fetch('/api/outreach-leads?id=' + encodeURIComponent(jobId), { cache: 'no-store' });
      if (!res.ok) {
        const j = await res.json().catch(() => ({}));
        throw new Error(j.error || ('HTTP ' + res.status));
      }
      viewModalState.sidecar = await res.json();
      renderViewModal();
    } catch (e) {
      $('view-modal-body').innerHTML = `<div class="modal-empty">⚠️ ${escapeHtml(e.message)}</div>`;
    }
  }

  function renderViewModal() {
    if (!viewModalState || !viewModalState.sidecar) return;
    const s = viewModalState.sidecar;
    const leads = s.leads || [];
    $('view-modal-title').textContent = `Leads — ${s.company || '(unknown)'}`;
    const stage = s.stage || 'unknown';
    $('view-modal-sub').textContent = `${leads.length} found · stage: ${stage}${s.brief_reason ? ' · ' + s.brief_reason : ''}`;

    const templates = s.templates || {};
    const tmplTypes = [
      { key: 'recruiter', label: 'Recruiter template', cls: 'tag-recruiter' },
      { key: 'founder',   label: 'Founder template',   cls: 'tag-founder' },
      { key: 'engineer',  label: 'Engineer template',  cls: 'tag-engineer' },
    ];

    let body = '';

    // Brief reason editor — affects per-lead message preview
    body += `<div class="brief-row"><label>Brief reason (what to highlight about this company):</label><input id="brief-reason-input" type="text" value="${escapeHtml(s.brief_reason || '')}" placeholder="e.g. ontology infra, AI dev tools, FDE program"></div>`;

    // Per-type templates with char counters
    for (const t of tmplTypes) {
      const val = templates[t.key] || '';
      const len = val.length;
      body += `
        <div class="tmpl-block">
          <div class="tmpl-head">
            <span><span class="lead-tag ${t.cls}">${t.key}</span> ${t.label}</span>
            <span class="char-counter ${len > 290 ? 'over' : ''}" id="tmpl-count-${t.key}">${len} chars</span>
          </div>
          <textarea data-tmpl="${t.key}" placeholder="Use {first_name}, {role_short}, {job_id_suffix}, {company}, {brief_reason} placeholders">${escapeHtml(val)}</textarea>
          <div class="tmpl-hint">Placeholders are hydrated per-lead. Edit a specific lead's message below to override for that one person.</div>
        </div>
      `;
    }

    if (!leads.length) {
      body += `<div class="modal-empty">No leads yet${stage === 'finding' ? ' — Claude is still searching…' : '. Click "Find more leads" to start.'}</div>`;
    } else {
      body += `<div class="lead-row-head"><div></div><div>Person</div><div>Message</div><div>Status</div><div></div></div>`;
      for (const L of leads) {
        const approved = L.approved === true;
        const status = (L.send_status || 'pending').replace('_', '-');
        const degTag = L.connection_degree === '1st' ? 'tag-1st' : (L.connection_degree === '2nd' ? 'tag-2nd' : 'tag-3rd');
        const typeTag = L.lead_type === 'recruiter' ? 'tag-recruiter' : (L.lead_type === 'founder' ? 'tag-founder' : 'tag-engineer');
        const msgLen = (L.message || '').length;
        const isSent = L.send_status === 'sent';
        body += `
          <div class="lead-row${approved ? '' : ' unapproved'}" data-lead-id="${escapeHtml(L.lead_id)}">
            <input type="checkbox" class="lead-cb" data-lead-id="${escapeHtml(L.lead_id)}" ${approved ? 'checked' : ''} ${isSent ? 'disabled' : ''}>
            <div class="lead-info">
              <div class="lead-name">${escapeHtml(L.name || '(no name)')}</div>
              <div class="lead-title">${escapeHtml(L.title || '')}</div>
              <div class="lead-summary">${escapeHtml(L.summary || '')}</div>
              <div class="lead-tags">
                <span class="lead-tag ${degTag}">${escapeHtml(L.connection_degree || '?')}</span>
                <span class="lead-tag ${typeTag}">${escapeHtml(L.lead_type || '?')}</span>
              </div>
              ${L.qualified_reason ? `<div style="font-size:10px;color:#94a3b8;margin-top:4px;font-style:italic">${escapeHtml(L.qualified_reason)}</div>` : ''}
            </div>
            <div class="lead-msg">
              <textarea data-lead-msg-id="${escapeHtml(L.lead_id)}" ${isSent ? 'disabled' : ''}>${escapeHtml(L.message || '')}</textarea>
              <div class="char-counter ${msgLen > 290 ? 'over' : ''}" id="msg-count-${escapeHtml(L.lead_id)}" style="margin-top:4px">${msgLen}/300</div>
            </div>
            <div><span class="lead-status status-${status}">${escapeHtml(L.send_status || 'pending')}</span>${L.send_error ? `<div style="font-size:10px;color:#b91c1c;margin-top:4px">${escapeHtml(L.send_error)}</div>` : ''}</div>
            <a class="lead-link" href="${escapeHtml(L.profile_url || '#')}" target="_blank" rel="noopener">Profile ↗</a>
          </div>
        `;
      }
    }
    $('view-modal-body').innerHTML = body;
    updateApproveSummary();
  }

  function updateApproveSummary() {
    if (!viewModalState || !viewModalState.sidecar) return;
    const leads = viewModalState.sidecar.leads || [];
    const approved = leads.filter(L => L.approved === true).length;
    const pending = leads.filter(L => L.approved === true && (L.send_status === 'pending' || !L.send_status || L.send_status === 'failed')).length;
    $('view-approve-summary').textContent = `${approved} approved · ${pending} pending send`;
  }

  function markDirty() { if (viewModalState) viewModalState.dirty = true; }

  async function saveViewModal(silent) {
    if (!viewModalState || !viewModalState.sidecar) return;
    const s = viewModalState.sidecar;
    const body = {
      id: viewModalState.jobId,
      templates: s.templates || {},
      brief_reason: s.brief_reason || '',
      leads: (s.leads || []).map(L => ({ lead_id: L.lead_id, approved: L.approved, message: L.message })),
    };
    try {
      const res = await fetch('/api/outreach-leads-update', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(body),
      });
      const j = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(j.error || ('HTTP ' + res.status));
      viewModalState.dirty = false;
      if (!silent) showToast('Saved', false);
    } catch (e) {
      if (!silent) showToast('Save failed: ' + e.message, false, false);
    }
  }

  async function sendOutreach() {
    if (!viewModalState) return;
    if (viewModalState.dirty) await saveViewModal(true);
    const approvedCount = (viewModalState.sidecar.leads || []).filter(L =>
      L.approved === true && (L.send_status === 'pending' || !L.send_status || L.send_status === 'failed')
    ).length;
    if (!approvedCount) { showToast('Nothing to send — approve at least one lead first.', false, false); return; }
    if (!confirm(`Send LinkedIn messages to ${approvedCount} approved leads? Claude will open Terminal + Chrome MCP and process them one-by-one with 45-90s jitter between each.`)) return;
    try {
      const res = await fetch('/api/outreach-send', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify({ id: viewModalState.jobId }),
      });
      const j = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(j.error || ('HTTP ' + res.status));
      showToast(`Sending to ${j.approved_count} leads…`, false);
      closeModal('view-modal');
      await loadJobs(true);
      maybeStartPolling();
    } catch (e) {
      showToast('Send failed: ' + e.message, false, false);
    }
  }

  // Live updates inside view modal — checkbox toggles + textarea edits.
  document.addEventListener('input', (e) => {
    if (!viewModalState || !viewModalState.sidecar) return;
    if (e.target.matches('textarea[data-tmpl]')) {
      const k = e.target.dataset.tmpl;
      viewModalState.sidecar.templates = viewModalState.sidecar.templates || {};
      viewModalState.sidecar.templates[k] = e.target.value;
      const len = e.target.value.length;
      const counter = document.getElementById('tmpl-count-' + k);
      if (counter) { counter.textContent = `${len} chars`; counter.classList.toggle('over', len > 290); }
      markDirty();
    } else if (e.target.matches('textarea[data-lead-msg-id]')) {
      const lid = e.target.dataset.leadMsgId;
      const lead = (viewModalState.sidecar.leads || []).find(L => L.lead_id === lid);
      if (lead) { lead.message = e.target.value; markDirty(); }
      const counter = document.getElementById('msg-count-' + lid);
      if (counter) { counter.textContent = `${e.target.value.length}/300`; counter.classList.toggle('over', e.target.value.length > 290); }
    } else if (e.target.id === 'brief-reason-input') {
      viewModalState.sidecar.brief_reason = e.target.value;
      markDirty();
    }
  });
  document.addEventListener('change', (e) => {
    if (!viewModalState || !viewModalState.sidecar) return;
    if (e.target.matches('input.lead-cb')) {
      const lid = e.target.dataset.leadId;
      const lead = (viewModalState.sidecar.leads || []).find(L => L.lead_id === lid);
      if (lead) {
        lead.approved = !!e.target.checked;
        markDirty();
        const row = e.target.closest('.lead-row');
        if (row) row.classList.toggle('unapproved', !lead.approved);
        updateApproveSummary();
      }
    }
  });

  document.addEventListener('click', (e) => {
    if (e.target.id === 'add-toggle') {
      const form = $('add-form');
      form.style.display = form.style.display === 'none' ? 'flex' : 'none';
      if (form.style.display === 'flex') $('add-jd-input').focus();
      return;
    }
    if (e.target.id === 'add-cancel') {
      $('add-form').style.display = 'none';
      return;
    }
    if (e.target.id === 'add-submit') {
      submitManualAdd();
      return;
    }
    // Outreach modal interactions
    if (e.target.matches('[data-close]')) { closeModal(e.target.dataset.close); return; }
    if (e.target.classList.contains('modal-bg')) { closeModal(e.target.id); return; }
    if (e.target.matches('[data-find-count]')) {
      const count = parseInt(e.target.dataset.findCount, 10);
      if (findModalJobId) triggerFind(findModalJobId, count, false);
      return;
    }
    if (e.target.id === 'view-find-more') {
      if (viewModalState) {
        if (confirm('Find more leads to append? Default: 10. Click cancel to enter custom.')) {
          triggerFind(viewModalState.jobId, 10, true);
        }
      }
      return;
    }
    if (e.target.id === 'view-approve-all') {
      if (!viewModalState) return;
      (viewModalState.sidecar.leads || []).forEach(L => {
        if (L.send_status !== 'sent') L.approved = true;
      });
      markDirty();
      renderViewModal();
      return;
    }
    if (e.target.id === 'view-unapprove-all') {
      if (!viewModalState) return;
      (viewModalState.sidecar.leads || []).forEach(L => {
        if (L.send_status !== 'sent') L.approved = false;
      });
      markDirty();
      renderViewModal();
      return;
    }
    if (e.target.id === 'view-save') { saveViewModal(false); return; }
    if (e.target.id === 'view-send') { sendOutreach(); return; }

    const chip = e.target.closest('.chip:not(#add-toggle)');
    if (chip) {
      activeFilter = chip.dataset.filter;
      localStorage.setItem('jobs-triage-filter', activeFilter);
      render();
      return;
    }
    const btn = e.target.closest('button[data-act]');
    if (btn) {
      const act = btn.dataset.act;
      if (act === 'outreach-find') { openFindModal(btn.dataset.id); return; }
      if (act === 'outreach-view') { openViewModal(btn.dataset.id); return; }
      decide(btn.dataset.id, act);
      return;
    }
    if (e.target.id === 'toast-undo') { doUndo(); $('toast').classList.add('hidden'); }
    if (e.target.id === 'reload-btn') { loadJobs(); }
    if (e.target.id === 'kill-all-btn') { killTerminals('all'); return; }
    // Row menu trigger
    const menuTrigger = e.target.closest('button[data-act="row-menu"]');
    if (menuTrigger) {
      e.stopPropagation();
      toggleRowMenu(menuTrigger.dataset.id, menuTrigger);
      return;
    }
    // Row menu items
    const menuItem = e.target.closest('button[data-row-act]');
    if (menuItem) {
      e.stopPropagation();
      if (menuItem.dataset.rowAct === 'kill') {
        killTerminals('job', menuItem.dataset.id);
      }
      return;
    }
    // Close any open row menu on outside click
    if (!e.target.closest('.row-menu')) closeAllRowMenus();
  });

  async function submitManualAdd() {
    const raw = $('add-jd-input').value.trim();
    if (!raw) { showToast('Paste a URL or JD text', false, false); return; }
    // Auto-detect: anything starting with http(s):// is a URL; everything else is JD text.
    const isUrl = /^https?:\/\//i.test(raw);
    const payload = isUrl ? { apply_url: raw } : { jd_text: raw };
    const submitBtn = $('add-submit');
    submitBtn.disabled = true;
    submitBtn.textContent = 'Saving…';
    try {
      const res = await fetch('/api/add-manual', {
        method: 'POST',
        headers: { 'Content-Type': 'application/json' },
        body: JSON.stringify(payload),
      });
      const json = await res.json().catch(() => ({}));
      if (!res.ok) throw new Error(json.error || ('HTTP ' + res.status));
      showToast(`#${json.id} saved · Claude is ${isUrl ? 'fetching the URL' : 'reading the JD'} and tailoring`, false);
      $('add-jd-input').value = '';
      $('add-form').style.display = 'none';
      activeFilter = 'pending';
      localStorage.setItem('jobs-triage-filter', activeFilter);
      await loadJobs(true);
    } catch (e) {
      showToast('Failed: ' + e.message, false, false);
    } finally {
      submitBtn.disabled = false;
      submitBtn.textContent = 'Add & Start Tailoring';
    }
  }

  loadJobs();
})();
</script>
</body>
</html>
"""


RUNS_HTML = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<title>📡 Live runs</title>
<link rel="icon" type="image/svg+xml" href="data:image/svg+xml,<svg xmlns='http://www.w3.org/2000/svg' viewBox='0 0 100 100'><text y='.9em' font-size='90'>%F0%9F%93%A1</text></svg>">
<style>
:root { color-scheme: light; }
* { box-sizing: border-box; }
html, body { margin:0; padding:0; background:#fafbfc; color:#0f172a; font-family:-apple-system,BlinkMacSystemFont,"Inter","Segoe UI",system-ui,sans-serif; font-size:14px; }
.top { display:flex; align-items:center; justify-content:space-between; padding:14px 18px; border-bottom:1px solid #e2e8f0; background:#fff; position:sticky; top:0; z-index:5; }
.title { font-size:17px; font-weight:650; }
.back-link { font-size:13px; color:#475569; text-decoration:none; padding:6px 12px; border:1px solid #e2e8f0; border-radius:6px; background:#fff; }
.back-link:hover { background:#f1f5f9; }
.wrap { display:flex; gap:0; height:calc(100vh - 57px); }
.side { width:300px; min-width:240px; border-right:1px solid #e2e8f0; overflow-y:auto; background:#fff; }
.run-item { padding:11px 14px; border-bottom:1px solid #f1f5f9; cursor:pointer; }
.run-item:hover { background:#f8fafc; }
.run-item.sel { background:#eef2ff; border-left:3px solid #6366f1; padding-left:11px; }
.run-label { font-weight:600; font-size:13px; }
.run-meta { font-size:11.5px; color:#64748b; margin-top:3px; display:flex; gap:8px; align-items:center; }
.pill { display:inline-flex; align-items:center; gap:5px; padding:1px 8px; border-radius:999px; font-size:10.5px; font-weight:700; }
.pill.live { background:#dbeafe; color:#1d4ed8; } .pill.live .dot { animation:pulse 1.2s infinite; }
.pill.done { background:#dcfce7; color:#15803d; }
.pill.idle { background:#f1f5f9; color:#64748b; }
.pill .dot { width:6px; height:6px; border-radius:50%; background:currentColor; }
@keyframes pulse { 0%,100%{opacity:1} 50%{opacity:.25} }
.feed { flex:1; overflow-y:auto; padding:16px 22px 80px; background:#0f172a; }
.empty { color:#64748b; padding:40px; text-align:center; }
.ev { margin:0 0 6px; max-width:980px; font-size:12.5px; line-height:1.5; }
.ev.text { background:#1e293b; color:#e2e8f0; border-radius:8px; padding:9px 13px; white-space:pre-wrap; border-left:3px solid #6366f1; }
.ev.tool { color:#7dd3fc; font-family:ui-monospace,Menlo,monospace; font-size:12px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; }
.ev.tool .tname { color:#38bdf8; font-weight:700; }
.ev.tres { color:#475569; font-family:ui-monospace,Menlo,monospace; font-size:11px; white-space:nowrap; overflow:hidden; text-overflow:ellipsis; padding-left:18px; }
.ev.tres.err { color:#f87171; }
.ev.info { color:#64748b; font-size:11.5px; }
.ev.raw { color:#475569; font-family:ui-monospace,Menlo,monospace; font-size:11px; white-space:pre-wrap; }
.ev.done { background:#064e3b; color:#d1fae5; border-radius:8px; padding:10px 14px; white-space:pre-wrap; border-left:3px solid #10b981; }
.ev.done.fail { background:#450a0a; color:#fecaca; border-left-color:#ef4444; }
.feed-head { color:#94a3b8; font-size:12px; padding-bottom:10px; border-bottom:1px solid #1e293b; margin-bottom:12px; }
</style>
</head>
<body>
<div class="top">
  <div class="title">📡 Live runs</div>
  <a class="back-link" href="/">← back to triage</a>
</div>
<div class="wrap">
  <div class="side" id="side"><div class="empty">loading…</div></div>
  <div class="feed" id="feed"><div class="empty">select a run on the left</div></div>
</div>
<script>
let selected = null, offset = 0, evTimer = null;
const TOOL_ICONS = { navigate:"🌐", computer:"🖱️", read_page:"📖", get_page_text:"📖", find:"🔍",
  Bash:"🔧", Read:"📄", Edit:"✏️", Write:"✏️", Glob:"🔍", Grep:"🔍", WebFetch:"🌐", WebSearch:"🔎", TodoWrite:"📝" };
function esc(s){ const d=document.createElement('div'); d.textContent=s||''; return d.innerHTML; }
function fmtTime(ts){ return new Date(ts*1000).toLocaleString([], {month:'short',day:'numeric',hour:'2-digit',minute:'2-digit'}); }

async function refreshList(){
  try {
    const r = await fetch('/api/runs-list'); const data = await r.json();
    const side = document.getElementById('side');
    if (!data.runs.length) { side.innerHTML = '<div class="empty">no runs in the last 48h</div>'; return; }
    side.innerHTML = data.runs.map(run => {
      const pill = run.active ? '<span class="pill live"><span class="dot"></span>LIVE</span>'
                 : run.done   ? '<span class="pill done">done</span>'
                              : '<span class="pill idle">idle</span>';
      return `<div class="run-item ${run.name===selected?'sel':''}" data-name="${esc(run.name)}">
        <div class="run-label">${esc(run.label)}</div>
        <div class="run-meta">${pill}<span>${fmtTime(run.mtime)}</span><span>${(run.size/1024).toFixed(0)}KB</span></div>
      </div>`;
    }).join('');
    side.querySelectorAll('.run-item').forEach(el =>
      el.addEventListener('click', () => selectRun(el.dataset.name)));
    // auto-select the newest LIVE run on first load
    if (!selected) { const live = data.runs.find(r=>r.active) || data.runs[0]; if (live) selectRun(live.name); }
  } catch(e) {}
}

function selectRun(name){
  selected = name; offset = 0;
  const feed = document.getElementById('feed');
  feed.innerHTML = `<div class="feed-head">${esc(name)}</div>`;
  document.querySelectorAll('.run-item').forEach(el => el.classList.toggle('sel', el.dataset.name===name));
  if (evTimer) clearInterval(evTimer);
  pollEvents(); evTimer = setInterval(pollEvents, 2000);
}

function render(ev){
  if (ev.t==='text') return `<div class="ev text">${esc(ev.text)}</div>`;
  if (ev.t==='tool') { const ic = TOOL_ICONS[ev.name] || (ev.name.startsWith('mcp__')?'🌐':'🔧');
    return `<div class="ev tool">${ic} <span class="tname">${esc(ev.name.replace('mcp__claude-in-chrome__',''))}</span> ${esc(ev.input)}</div>`; }
  if (ev.t==='tool_result') return ev.text ? `<div class="ev tres ${ev.is_error?'err':''}">↳ ${esc(ev.text)}</div>` : '';
  if (ev.t==='info') return `<div class="ev info">ⓘ ${esc(ev.text)}</div>`;
  if (ev.t==='done') { const mins = Math.floor(ev.duration_s/60), secs = ev.duration_s%60;
    const cost = ev.cost ? ` · $${ev.cost.toFixed(2)}` : '';
    return `<div class="ev done ${ev.ok?'':'fail'}">${ev.ok?'✅':'❌'} finished in ${mins}m ${secs}s${cost}\n${esc(ev.text)}</div>`; }
  if (ev.t==='raw') return `<div class="ev raw">${esc(ev.text)}</div>`;
  return '';
}

async function pollEvents(){
  if (!selected) return;
  try {
    const r = await fetch(`/api/run-events?name=${encodeURIComponent(selected)}&offset=${offset}`);
    if (!r.ok) return;
    const data = await r.json();
    offset = data.offset;
    if (data.events.length) {
      const feed = document.getElementById('feed');
      const pinned = feed.scrollHeight - feed.scrollTop - feed.clientHeight < 120;
      feed.insertAdjacentHTML('beforeend', data.events.map(render).join(''));
      if (pinned) feed.scrollTop = feed.scrollHeight;
    }
  } catch(e) {}
}

refreshList(); setInterval(refreshList, 5000);
</script>
</body>
</html>
"""


# ---------------------------------------------------------------------------
# HTTP server
# ---------------------------------------------------------------------------

class Handler(BaseHTTPRequestHandler):
    def log_message(self, fmt, *args):
        sys.stderr.write("[%s] %s\n" % (self.log_date_time_string(), fmt % args))

    def _send(self, code, body, content_type):
        if isinstance(body, str):
            body = body.encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(body)))
        self.send_header("Cache-Control", "no-store")
        self.end_headers()
        self.wfile.write(body)

    def _send_json(self, code, data):
        self._send(code, json.dumps(data, default=str), "application/json; charset=utf-8")

    def do_GET(self):
        if self.path in ("/", "/index.html"):
            self._send(200, HTML, "text/html; charset=utf-8")
            return
        if self.path in ("/runs", "/runs/"):
            self._send(200, RUNS_HTML, "text/html; charset=utf-8")
            return
        if self.path == "/api/jobs":
            self._handle_jobs()
            return
        if self.path == "/api/run-status":
            self._handle_run_status()
            return
        if self.path == "/api/runs-list":
            self._handle_runs_list()
            return
        if self.path.startswith("/api/run-events"):
            self._handle_run_events()
            return
        if self.path == "/api/health":
            self._send_json(200, {"ok": True, "xlsx": str(XLSX_PATH), "xlsx_exists": XLSX_PATH.exists(), "claude_bin": CLAUDE_BIN})
            return
        if self.path.startswith("/api/outreach-leads"):
            self._handle_outreach_leads_get()
            return
        if self.path.startswith("/resume-files/"):
            self._serve_resume_file(self.path[len("/resume-files/"):])
            return
        self._send(404, "Not Found", "text/plain")

    RUN_LOG_KINDS = [
        ("sourcing-", "sourcing", "🌐 Daily sourcing"),
        ("tailor-", "tailor", "✂️ Resume tailor"),
        ("outreach-find-", "outreach-find", "🔎 Outreach: find leads"),
        ("outreach-send-", "outreach-send", "📨 Outreach: send"),
        ("verify-date-", "verify-date", "📅 Verify date"),
    ]
    RUN_LOG_NAME_RE = re.compile(r"^[A-Za-z0-9._-]+\.log$")

    def _handle_runs_list(self):
        """GET /api/runs-list — recent spawned runs with live/done state."""
        now = datetime.datetime.now().timestamp()
        runs = []
        for p in LOGS_DIR.glob("*.log"):
            kind = label = None
            for prefix, k, lab in self.RUN_LOG_KINDS:
                if p.name.startswith(prefix):
                    kind, label = k, lab
                    break
            if not kind:
                continue
            mtime = p.stat().st_mtime
            if now - mtime > 48 * 3600:
                continue
            # a finished stream-json run always ends with a {"type":"result"} event
            done = False
            try:
                with open(p, "rb") as fh:
                    fh.seek(max(0, p.stat().st_size - 8192))
                    tail = fh.read()
                    done = b'"type":"result"' in tail or b'"type": "result"' in tail
            except OSError:
                pass
            runs.append({
                "name": p.name,
                "kind": kind,
                "label": label,
                "mtime": int(mtime),
                "size": p.stat().st_size,
                "active": (now - mtime) < 90 and not done,
                "done": done,
            })
        runs.sort(key=lambda r: r["mtime"], reverse=True)
        self._send_json(200, {"runs": runs[:40]})

    @staticmethod
    def _simplify_stream_event(obj):
        """Map a claude stream-json event to a small renderable dict."""
        t = obj.get("type")
        if t == "system" and obj.get("subtype") == "init":
            return {"t": "info", "text": f"session started · model {obj.get('model', '?')}"}
        if t == "assistant":
            out = []
            for block in (obj.get("message") or {}).get("content") or []:
                bt = block.get("type")
                if bt == "text" and (block.get("text") or "").strip():
                    out.append({"t": "text", "text": block["text"][:800]})
                elif bt == "tool_use":
                    inp = json.dumps(block.get("input") or {}, ensure_ascii=False)
                    out.append({"t": "tool", "name": block.get("name", "?"),
                                "input": inp[:300]})
            return out or None
        if t == "user":
            out = []
            for block in (obj.get("message") or {}).get("content") or []:
                if isinstance(block, dict) and block.get("type") == "tool_result":
                    content = block.get("content")
                    if isinstance(content, list):
                        content = " ".join(
                            c.get("text", "") for c in content if isinstance(c, dict))
                    text = (content or "")[:400] if isinstance(content, str) else ""
                    out.append({"t": "tool_result", "text": text,
                                "is_error": bool(block.get("is_error"))})
            return out or None
        if t == "result":
            return {"t": "done",
                    "ok": obj.get("subtype") == "success",
                    "duration_s": round((obj.get("duration_ms") or 0) / 1000),
                    "cost": obj.get("total_cost_usd"),
                    "text": (obj.get("result") or "")[:1500]}
        return None

    def _handle_run_events(self):
        """GET /api/run-events?name=<log>&offset=<bytes> — incremental tail."""
        from urllib.parse import urlparse, parse_qs
        q = parse_qs(urlparse(self.path).query)
        name = (q.get("name") or [""])[0]
        offset = int((q.get("offset") or ["0"])[0])
        if not self.RUN_LOG_NAME_RE.match(name):
            self._send_json(400, {"error": "bad name"})
            return
        path = LOGS_DIR / name
        if not path.exists() or path.resolve().parent != LOGS_DIR.resolve():
            self._send_json(404, {"error": "no such log"})
            return
        size = path.stat().st_size
        if offset >= size:
            self._send_json(200, {"events": [], "offset": offset,
                                  "active": _is_run_active(path)})
            return
        with open(path, "rb") as fh:
            fh.seek(offset)
            chunk = fh.read(min(size - offset, 2_000_000))
        # only consume up to the last full line; a partial trailing line waits
        last_nl = chunk.rfind(b"\n")
        if last_nl < 0:
            self._send_json(200, {"events": [], "offset": offset,
                                  "active": _is_run_active(path)})
            return
        consumed = chunk[: last_nl + 1]
        events = []
        for raw in consumed.decode("utf-8", errors="replace").splitlines():
            raw = raw.strip()
            if not raw:
                continue
            try:
                simplified = self._simplify_stream_event(json.loads(raw))
            except (json.JSONDecodeError, AttributeError):
                simplified = {"t": "raw", "text": _strip_ansi(raw)[:400]}
            if simplified is None:
                continue
            if isinstance(simplified, list):
                events.extend(simplified)
            else:
                events.append(simplified)
        self._send_json(200, {"events": events[:500],
                              "offset": offset + len(consumed),
                              "active": _is_run_active(path)})

    def _handle_run_status(self):
        """Return the active sourcing run's stage + log tail. Used by /runs."""
        try:
            log_path = _find_active_sourcing_log()
            if not log_path:
                self._send_json(200, {
                    "has_log": False,
                    "is_active": False,
                    "stages": [{"id": s, "name": n, "state": "pending"} for s, n, _ in STAGE_DEFS],
                    "tail": "",
                    "log_path": None,
                    "log_mtime": None,
                })
                return

            text = log_path.read_text(encoding="utf-8", errors="replace")
            stripped = _strip_ansi(text)
            current, completed = _detect_stages(stripped)
            active = _is_run_active(log_path)

            stages = []
            for sid, name, _ in STAGE_DEFS:
                if sid == current:
                    state = "active" if active else "stalled"
                elif sid in completed:
                    state = "completed"
                else:
                    state = "pending"
                stages.append({"id": sid, "name": name, "state": state})

            # Tail: last ~60 cleaned, non-empty lines
            lines = [ln.rstrip() for ln in stripped.splitlines() if ln.strip()]
            tail = "\n".join(lines[-60:])

            self._send_json(200, {
                "has_log": True,
                "is_active": active,
                "stages": stages,
                "tail": tail,
                "log_path": str(log_path),
                "log_mtime": datetime.datetime.fromtimestamp(log_path.stat().st_mtime).isoformat(),
                "current_stage": current,
            })
        except Exception as e:
            self._send_json(500, {"error": str(e)})

    def _serve_resume_file(self, relpath: str):
        """Serve files from RESUME_ROOT — needed so the UI can link to tailored PDFs."""
        try:
            relpath = relpath.split("?", 1)[0].split("#", 1)[0]
            # URL-decode (filenames have spaces per the resume.filename_pattern config)
            relpath = unquote(relpath)
            target = (RESUME_ROOT / relpath).resolve()
            # Safety: must be inside RESUME_ROOT
            if RESUME_ROOT.resolve() not in target.parents and target != RESUME_ROOT.resolve():
                self._send(403, "Forbidden", "text/plain")
                return
            if not target.exists() or not target.is_file():
                self._send(404, "Not Found", "text/plain")
                return
            ctype, _ = mimetypes.guess_type(str(target))
            ctype = ctype or "application/octet-stream"
            data = target.read_bytes()
            self.send_response(200)
            self.send_header("Content-Type", ctype)
            self.send_header("Content-Length", str(len(data)))
            self.send_header("Cache-Control", "no-store")
            self.end_headers()
            self.wfile.write(data)
        except Exception as e:
            self._send(500, f"Error: {e}", "text/plain")

    def do_POST(self):
        if self.path == "/api/decide":
            self._handle_decide()
            return
        if self.path == "/api/cancel":
            self._handle_cancel()
            return
        if self.path == "/api/add-manual":
            self._handle_add_manual()
            return
        if self.path == "/api/applied":
            self._handle_applied()
            return
        if self.path == "/api/outreach-find":
            self._handle_outreach_find()
            return
        if self.path == "/api/outreach-leads-update":
            self._handle_outreach_leads_update()
            return
        if self.path == "/api/outreach-send":
            self._handle_outreach_send()
            return
        if self.path == "/api/verify-date":
            self._handle_verify_date()
            return
        if self.path == "/api/kill-terminals":
            self._handle_kill_terminals()
            return
        self._send(404, "Not Found", "text/plain")

    def _handle_jobs(self):
        try:
            if not XLSX_PATH.exists():
                self._send_json(500, {"error": f"xlsx not found at {XLSX_PATH}"})
                return
            wb = load_workbook(XLSX_PATH, data_only=True)
            ws = wb["jobs"]
            headers = [c.value for c in ws[1]]
            rows = []
            counts = {"pending": 0, "yes": 0, "no": 0, "maybe": 0}
            for r in ws.iter_rows(min_row=2, values_only=True):
                if not any(c not in (None, "") for c in r):
                    continue
                row = {k: ("" if v is None else str(v)) for k, v in zip(headers, r)}
                decision = (row.get("decision") or "").lower()
                counts[decision] = counts.get(decision, 0) + 1
                # Inject outreach summary if a sidecar exists. Cheap: only fires
                # for rows that have an applied_date (the only ones the UI shows
                # the Outreach button on).
                if row.get("applied_date"):
                    try:
                        rid = int(row.get("id") or 0)
                        if rid:
                            row.update(_outreach_summary(rid))
                    except (ValueError, TypeError):
                        pass
                rows.append(row)
            self._send_json(200, {"rows": rows, "stats": counts})
        except Exception as e:
            self._send_json(500, {"error": str(e)})

    def _handle_decide(self):
        try:
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length).decode("utf-8")
            data = json.loads(body) if body else {}
            job_id = int(data["id"])
            decision = str(data["decision"]).strip().lower()
            if decision not in {"pending", "yes", "no", "maybe"}:
                self._send_json(400, {"error": "decision must be pending/yes/no/maybe"})
                return
            try:
                wb = load_workbook(XLSX_PATH)
            except PermissionError:
                self._send_json(423, {"error": "jobs.xlsx is locked — close Excel first"})
                return
            ws = wb["jobs"]
            headers = [c.value for c in ws[1]]
            if "decision" not in headers:
                self._send_json(500, {"error": "no 'decision' column"})
                return
            decision_idx = headers.index("decision")
            rv_idx = headers.index("resume_version") if "resume_version" in headers else None

            # Parallel tailoring is allowed — clicking ✓ Yes on multiple jobs spawns a
            # separate Terminal window per job, each running its own Claude Code session.
            # The xlsx write at completion uses openpyxl (load+save). If two tailors
            # finish within ~1 second of each other, the later save wins; acceptable
            # tradeoff for parallelism. The PDFs themselves are written to distinct
            # paths per (company, role) slug so they never collide on disk.

            found = False
            prev_resume_version = ""
            should_tailor = False
            for r in ws.iter_rows(min_row=2):
                if r[0].value == job_id:
                    r[decision_idx].value = decision
                    if rv_idx is not None:
                        prev_resume_version = (r[rv_idx].value or "")
                        # Only trigger tailoring when transitioning to yes from a non-yes / un-tailored state
                        if decision == "yes" and (not prev_resume_version or prev_resume_version.startswith("error:")):
                            r[rv_idx].value = "tailoring"
                            should_tailor = True
                    found = True
                    break
            if not found:
                self._send_json(404, {"error": f"id {job_id} not found"})
                return
            wb.save(XLSX_PATH)

            # Spawn Claude Code tailor pass AFTER the xlsx is saved
            log_path = None
            if should_tailor:
                log_path = spawn_claude_tailor(job_id)

            self._send_json(200, {
                "ok": True,
                "id": job_id,
                "decision": decision,
                "tailoring_started": should_tailor,
                "tailor_log": str(log_path) if log_path else None,
            })
        except Exception as e:
            self._send_json(400, {"error": str(e)})

    def _handle_add_manual(self):
        """Add a manually-pasted job row + immediately spawn the tailoring session.

        Body: { company, role, location?, apply_url?, comp?, jd_text }
        Returns: { ok: true, id: <new_id>, tailor_log: <path> }

        The row is inserted with decision=yes and resume_version=tailoring so the
        existing card UI shows the spinner + Cancel button immediately. date_sourced
        is set to today (the date the row was saved). source=manual, tier=manual.
        The jd_text field is added to the schema on first use if missing — the
        tailoring prompt instructs Claude to use that field instead of web_fetching.
        """
        try:
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length).decode("utf-8")
            data = json.loads(body) if body else {}
            jd_text = (data.get("jd_text") or "").strip()
            apply_url = (data.get("apply_url") or "").strip()
            if not jd_text and not apply_url:
                self._send_json(400, {"error": "paste either a job URL or the full JD text"})
                return
            # All other fields are optional — Claude extracts them from the URL or JD
            # during the first step of the tailoring run. Placeholder values let the
            # card render immediately so the user sees the row land in the pending list.
            company = (data.get("company") or "").strip() or "(extracting…)"
            role = (data.get("role") or "").strip() or "(extracting…)"
            location = (data.get("location") or "").strip()
            comp = (data.get("comp") or "").strip()

            try:
                wb = load_workbook(XLSX_PATH)
            except PermissionError:
                self._send_json(423, {"error": "jobs.xlsx is locked — close Excel first"})
                return
            ws = wb["jobs"]
            headers = [c.value for c in ws[1]]

            # Schema migration: add jd_text column on first use if missing.
            # NOTE: don't try to copy .font from another header cell — openpyxl returns
            # StyleProxy objects that can't be reassigned and cause unhashable-type errors
            # on wb.save(). The plain header value is fine; Excel will render it normally.
            if "jd_text" not in headers:
                ws.cell(row=1, column=len(headers) + 1, value="jd_text")
                headers.append("jd_text")

            # Compute next id.
            max_id = 0
            for r in ws.iter_rows(min_row=2, values_only=True):
                try:
                    rid = int(r[0]) if r[0] is not None else 0
                    if rid > max_id:
                        max_id = rid
                except (ValueError, TypeError):
                    pass
            new_id = max_id + 1

            today = datetime.date.today().isoformat()
            row_data = {
                "id": new_id,
                "date_sourced": today,
                "posted_date": today,
                "freshness": "manual",
                "company": company,
                "role": role,
                "location": location,
                "comp": comp,
                "source": "manual",
                "apply_url": apply_url,
                "tier": "manual",
                "reasoning": "manually pasted by user via triage UI",
                "decision": "yes",
                "applied_date": "",
                "resume_version": "tailoring",
                "notes": "",
                "jd_text": jd_text,
            }
            row_values = [row_data.get(h, "") for h in headers]
            ws.append(row_values)
            wb.save(XLSX_PATH)

            # Fire tailoring in a background thread so the HTTP response doesn't block
            # on AppleScript (which can take a few seconds). The card will reflect the
            # tailoring state from the xlsx on the next poll.
            log_path = LOGS_DIR / f"tailor-{new_id}-deferred.log"
            threading.Thread(
                target=spawn_claude_tailor,
                args=(new_id,),
                daemon=True,
            ).start()

            self._send_json(200, {
                "ok": True,
                "id": new_id,
                "company": company,
                "role": role,
                "date_sourced": today,
                "tailor_log": str(log_path),
            })
        except Exception as e:
            self._send_json(400, {"error": str(e)})

    def _handle_applied(self):
        """Toggle the applied_date column for a row.

        Body: { id: int, applied_date?: str }
          - applied_date omitted or null → set to today's ISO date.
          - applied_date == "" → CLEAR the cell (un-apply).
          - applied_date == "YYYY-MM-DD" → set to that explicit date.
        """
        try:
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length).decode("utf-8")
            data = json.loads(body) if body else {}
            job_id = int(data["id"])
            if "applied_date" in data and data["applied_date"] is not None:
                new_val = str(data["applied_date"]).strip()
            else:
                new_val = datetime.date.today().isoformat()
            try:
                wb = load_workbook(XLSX_PATH)
            except PermissionError:
                self._send_json(423, {"error": "jobs.xlsx is locked — close Excel first"})
                return
            ws = wb["jobs"]
            headers = [c.value for c in ws[1]]
            if "applied_date" not in headers:
                self._send_json(500, {"error": "no 'applied_date' column"})
                return
            ad_idx = headers.index("applied_date")
            found = False
            for r in ws.iter_rows(min_row=2):
                if r[0].value == job_id:
                    r[ad_idx].value = new_val
                    found = True
                    break
            if not found:
                self._send_json(404, {"error": f"id {job_id} not found"})
                return
            wb.save(XLSX_PATH)
            self._send_json(200, {"ok": True, "id": job_id, "applied_date": new_val})
        except Exception as e:
            self._send_json(400, {"error": str(e)})

    def _handle_outreach_leads_get(self):
        """GET /api/outreach-leads?id=<job_id> — return the sidecar JSON, or 404."""
        try:
            from urllib.parse import urlparse, parse_qs
            qs = parse_qs(urlparse(self.path).query)
            job_id_raw = (qs.get("id") or [""])[0]
            job_id = int(job_id_raw)
            data = _load_sidecar(job_id)
            if data is None:
                self._send_json(404, {"error": f"no sidecar for job_id={job_id}"})
                return
            self._send_json(200, data)
        except Exception as e:
            self._send_json(400, {"error": str(e)})

    def _handle_outreach_find(self):
        """POST /api/outreach-find — spawn Stage 1 finder.

        Body: { id: int, count: int, append?: bool }
        Returns: { ok: true, log: <path> }
        """
        try:
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length).decode("utf-8")
            data = json.loads(body) if body else {}
            job_id = int(data["id"])
            lead_count = int(data.get("count", 10))
            if lead_count < 1 or lead_count > 30:
                self._send_json(400, {"error": "count must be 1..30"})
                return
            append = bool(data.get("append", False))

            # If append=false and a sidecar already exists with leads, prevent
            # accidental overwrite. UI should call with append=true in that case.
            existing = _load_sidecar(job_id)
            if existing and (existing.get("leads") or []) and not append:
                self._send_json(409, {
                    "error": "sidecar already has leads — pass append=true to add more, or delete the sidecar to restart"
                })
                return

            # Seed a minimal sidecar so the UI can show "finding leads…" immediately,
            # before the spawned claude has done anything.
            if not existing:
                sidecar = _outreach_sidecar(job_id)
                sidecar.write_text(json.dumps({
                    "job_id": job_id,
                    "stage": "finding",
                    "leads": [],
                    "lead_count_requested": lead_count,
                }, indent=2))
            else:
                existing["stage"] = "finding"
                _outreach_sidecar(job_id).write_text(json.dumps(existing, indent=2))

            log_path = _spawn_claude_outreach(job_id, "find", lead_count=lead_count, append=append)
            self._send_json(200, {"ok": True, "id": job_id, "log": str(log_path)})
        except Exception as e:
            self._send_json(400, {"error": str(e)})

    def _handle_outreach_leads_update(self):
        """POST /api/outreach-leads-update — user toggles approve/edits messages.

        Body: {
          id: int,
          templates?: { recruiter, founder, engineer },
          brief_reason?: str,
          leads?: [ { lead_id, approved?, message? }, ... ]
        }
        Mutates only the fields passed; leaves everything else intact.
        """
        try:
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length).decode("utf-8")
            data = json.loads(body) if body else {}
            job_id = int(data["id"])
            sidecar_path = _outreach_sidecar(job_id)
            sidecar = _load_sidecar(job_id)
            if sidecar is None:
                self._send_json(404, {"error": f"no sidecar for job_id={job_id}"})
                return

            if "templates" in data and isinstance(data["templates"], dict):
                sidecar.setdefault("templates", {}).update(data["templates"])
            if "brief_reason" in data:
                sidecar["brief_reason"] = data["brief_reason"]
            if "leads" in data and isinstance(data["leads"], list):
                # Build a lookup of existing leads by lead_id
                idx = {L.get("lead_id"): L for L in sidecar.get("leads", []) if L.get("lead_id")}
                for patch in data["leads"]:
                    lid = patch.get("lead_id")
                    if not lid or lid not in idx:
                        continue
                    if "approved" in patch:
                        idx[lid]["approved"] = patch["approved"]
                    if "message" in patch:
                        idx[lid]["message"] = patch["message"]

            sidecar_path.write_text(json.dumps(sidecar, indent=2, ensure_ascii=False))
            self._send_json(200, {"ok": True, "id": job_id})
        except Exception as e:
            self._send_json(400, {"error": str(e)})

    def _handle_outreach_send(self):
        """POST /api/outreach-send — spawn Stage 2 sender for the approved subset.

        Body: { id: int }
        Returns: { ok: true, log: <path>, approved_count: N }

        Refuses if there are zero approved leads (nothing to send).
        """
        try:
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length).decode("utf-8")
            data = json.loads(body) if body else {}
            job_id = int(data["id"])
            sidecar = _load_sidecar(job_id)
            if sidecar is None:
                self._send_json(404, {"error": f"no sidecar for job_id={job_id}"})
                return
            approved = [
                L for L in (sidecar.get("leads") or [])
                if L.get("approved") is True
                and L.get("send_status") in (None, "pending", "failed")
            ]
            if not approved:
                self._send_json(400, {"error": "no approved leads pending send"})
                return
            log_path = _spawn_claude_outreach(job_id, "send")
            self._send_json(200, {"ok": True, "id": job_id, "log": str(log_path), "approved_count": len(approved)})
        except Exception as e:
            self._send_json(400, {"error": str(e)})

    def _handle_kill_terminals(self):
        """POST /api/kill-terminals — close Terminal.app windows + kill claude procs.

        Body: { kind: "all" | "job", id?: int }
          - "all": closes every Terminal window whose name starts with "claude-",
                  kills every spawn-claude (matched by spec file in cmdline),
                  resets every in-flight xlsx / sidecar field.
          - "job": same but scoped to a single job_id. Windows matched by
                  containing "-<id>-" in title; processes matched by JOB_ID/job_id
                  in cmdline; state reset for that one row + sidecar only.

        Never touches iTerm2 (script targets `application "Terminal"`).
        Never touches the user's interactive claude session in iTerm2 (those don't
        run any spec file from SPAWN_SPEC_PATTERNS, so pgrep doesn't match them).
        """
        try:
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length).decode("utf-8")
            data = json.loads(body) if body else {}
            kind = (data.get("kind") or "all").lower()
            if kind == "job":
                job_id = int(data["id"])
                # Window title pattern: claude-<task>-<id>-<ts>. We match by "-<id>-"
                # to catch any task type. Could match other rows with overlapping
                # substrings (e.g., job 1 matching job 15), so we use a more
                # specific filter that includes the dash separators.
                closed = 0
                for prefix in ("claude-tailor-", "claude-outreach-find-",
                              "claude-outreach-send-", "claude-verify-date-"):
                    closed += _close_terminal_windows(f"{prefix}{job_id}-")
                killed = _kill_orphan_claude_procs(job_id)
                reset = _reset_inflight_state(job_id)
                self._send_json(200, {
                    "ok": True, "kind": "job", "id": job_id,
                    "closed": closed, "killed": killed, "reset": reset,
                })
            else:
                # Kill all task-spawn claude terminals + procs + reset state.
                # Title prefix "claude-" matches all four task types + sourcing,
                # but the helper requires the substring to contain "claude-" so
                # it can never accidentally close non-claude Terminal windows.
                closed = _close_terminal_windows("claude-")
                killed = _kill_orphan_claude_procs(None)
                reset = _reset_inflight_state(None)
                self._send_json(200, {
                    "ok": True, "kind": "all",
                    "closed": closed, "killed": killed, "reset": reset,
                })
        except Exception as e:
            self._send_json(400, {"error": str(e)})

    def _handle_verify_date(self):
        """POST /api/verify-date — spawn Claude Code with Chrome to verify a job's real posted date.

        Body: { id: int }
        Returns: { ok: true, log: <path> }

        Optimistically marks the row's posted_date_verified column as "verifying"
        so the UI can show a spinner. Claude overwrites with YYYY-MM-DD or "unknown".
        """
        try:
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length).decode("utf-8")
            data = json.loads(body) if body else {}
            job_id = int(data["id"])
            if not _set_posted_date_verified(job_id, "verifying"):
                self._send_json(404, {"error": f"id {job_id} not found"})
                return
            log_path = _spawn_claude_verify_date(job_id)
            self._send_json(200, {"ok": True, "id": job_id, "log": str(log_path)})
        except Exception as e:
            self._send_json(400, {"error": str(e)})

    def _handle_cancel(self):
        """Cancel an in-flight tailoring: clear resume_version, revert decision to pending.

        The spawned Claude Code process (and its Terminal window) is NOT killed —
        The user closes that manually. This endpoint just unblocks the UI so the
        user can re-trigger or pick a different job.
        """
        try:
            length = int(self.headers.get("Content-Length") or 0)
            body = self.rfile.read(length).decode("utf-8")
            data = json.loads(body) if body else {}
            job_id = int(data["id"])
            try:
                wb = load_workbook(XLSX_PATH)
            except PermissionError:
                self._send_json(423, {"error": "jobs.xlsx is locked — close Excel first"})
                return
            ws = wb["jobs"]
            headers = [c.value for c in ws[1]]
            decision_idx = headers.index("decision") if "decision" in headers else None
            rv_idx = headers.index("resume_version") if "resume_version" in headers else None
            found = False
            for r in ws.iter_rows(min_row=2):
                if r[0].value == job_id:
                    if decision_idx is not None:
                        r[decision_idx].value = "pending"
                    if rv_idx is not None:
                        r[rv_idx].value = ""
                    found = True
                    break
            if not found:
                self._send_json(404, {"error": f"id {job_id} not found"})
                return
            wb.save(XLSX_PATH)
            self._send_json(200, {"ok": True, "id": job_id, "decision": "pending"})
        except Exception as e:
            self._send_json(400, {"error": str(e)})


def open_browser_soon():
    def _open():
        try:
            webbrowser.open(f"http://localhost:{PORT}/")
        except Exception:
            pass
    t = threading.Timer(0.8, _open)
    t.daemon = True
    t.start()


def main():
    print("=" * 60)
    print(f"  Jobs Triage Server")
    print("=" * 60)
    print(f"  xlsx file : {XLSX_PATH}")
    print(f"  url       : http://localhost:{PORT}/")
    print(f"  stop      : Ctrl+C")
    print("=" * 60)
    if not XLSX_PATH.exists():
        print(f"⚠️  WARNING: {XLSX_PATH} does not exist yet.")
        print("   Run the daily job sourcing pipeline first.\n")
    open_browser_soon()
    server = HTTPServer(("127.0.0.1", PORT), Handler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down.")
        server.server_close()


if __name__ == "__main__":
    main()
